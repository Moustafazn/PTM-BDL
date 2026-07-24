"""
HeLa / HDAC Inhibitor — Application-specific biological knowledge.

ALL application-specific labels, PTM sites, drug classifications, and
biological validation targets live here. The framework packages (ptm_bdl.*)
are protein-agnostic.

Biological context:
  HeLa — cervical carcinoma (HPV18+), standard epigenetics model cell line
    • 6 drugs: Vorinostat, Romidepsin, CUDC-101, A485, A486, Curcumin
    • PTM types: phosphorylation (S/T/Y) + acetylation (K) — NEW PTM type
    • 59K acetylation rows + 921K phosphorylation rows in DrugPTM-Bench

Target proteins for acetylation analysis:
  EP300 (P300 HAT, UniProt Q09472) — histone acetyltransferase, A485 target
  CREBBP (CBP, UniProt Q92793) — paralog of P300, HDAC substrate
  HDAC1 (UniProt Q13547) — primary target of Vorinostat/SAHA and Romidepsin
  HDAC2 (UniProt Q92769) — primary target, often in complex with HDAC1
  Histones (H3, H4) — downstream effectors of HDAC inhibition
"""

from __future__ import annotations

# ══════════════════════════════════════════════════════════════════════════════
# Drug Classification
# ══════════════════════════════════════════════════════════════════════════════

DRUGS = {
    "Vorinostat": {
        "mechanism": "Pan-HDAC inhibitor (SAHA)",
        "targets": ["HDAC1", "HDAC2", "HDAC3", "HDAC6"],
        "gdsc_id": 1012,
        "class": "HDAC inhibitor",
    },
    "Romidepsin": {
        "mechanism": "HDAC class I selective inhibitor (depsipeptide)",
        "targets": ["HDAC1", "HDAC2"],
        "gdsc_id": 1659,
        "class": "HDAC inhibitor",
    },
    "CUDC101": {
        "mechanism": "Triple HDAC/EGFR/HER2 inhibitor",
        "targets": ["HDAC1", "HDAC2", "HDAC3", "EGFR", "ERBB2"],
        "gdsc_id": 1578,
        "class": "HDAC/RTK inhibitor",
    },
    "A485": {
        "mechanism": "p300/CBP HAT inhibitor (catalytic)",
        "targets": ["EP300", "CREBBP"],
        "gdsc_id": None,
        "class": "HAT inhibitor",
    },
    "A486": {
        "mechanism": "Inactive control for A485",
        "targets": [],
        "gdsc_id": None,
        "class": "Negative control",
    },
    "Curcumin": {
        "mechanism": "Natural polyphenol — pleiotropic HDAC/HAT modulator",
        "targets": ["HDAC1", "HDAC3", "EP300"],
        "gdsc_id": None,
        "class": "Natural HDAC modulator",
    },
}

DRUGS_WITH_GDSC = ["Vorinostat", "Romidepsin", "CUDC101"]
DRUGS_WITHOUT_GDSC = ["A485", "A486", "Curcumin"]

# ══════════════════════════════════════════════════════════════════════════════
# Top Acetylation Genes (from DrugPTM-Bench HeLa data scan)
# ══════════════════════════════════════════════════════════════════════════════

TOP_ACETYLATION_GENES = [
    ("NCL", 1145),      # Nucleolin
    ("EP300", 948),      # p300 HAT — A485 target
    ("HIST1H4A", 716),   # Histone H4
    ("CREBBP", 674),     # CBP — p300 paralog
    ("NOLC1", 528),      # Nucleolar and coiled-body phosphoprotein
    ("NPM1", 501),       # Nucleophosmin
    ("H2AFZ;H2AFV", 475),  # Histone H2A variants
    ("ING4", 436),       # Inhibitor of growth 4
    ("HSPD1", 424),      # HSP60
    ("ENO1", 423),       # Enolase 1
]

# ══════════════════════════════════════════════════════════════════════════════
# HDAC Proteins in Data (from DrugPTM-Bench scan)
# ══════════════════════════════════════════════════════════════════════════════

HDAC_GENES_IN_DATA = {
    "HDAC1": 310,
    "HDAC2": 271,
    "HDAC3": 38,
    "HDAC4": 222,
    "HDAC5": 49,
    "HDAC6": 28,
    "HDAC7": 437,
}

# ══════════════════════════════════════════════════════════════════════════════
# Biological Validation Targets
# ══════════════════════════════════════════════════════════════════════════════

VALIDATION_TARGETS = {
    "acetyl_increases_under_hdac_inhibition": {
        "description": "HDAC inhibitors should INCREASE acetylation at histone marks",
        "expected": "Positive delta for histone Kac sites under Vorinostat/Romidepsin",
        "references": [
            "Hartl et al., Cell Reports 2024 — dose-resolved HDAC inhibitor proteomics",
            "Marks & Xu, J Cell Biochem 2009 (PMID 19479898)",
        ],
    },
    "a485_decreases_p300_acetylation": {
        "description": "A485 (HAT inhibitor) should DECREASE p300-mediated acetylation",
        "expected": "Negative delta for EP300 autoacetylation sites under A485",
        "references": [
            "Lasko et al., Nature 2017 (PMID 29211713) — A485/A486 mechanism",
        ],
    },
    "a486_no_effect": {
        "description": "A486 (inactive control) should show minimal PTM changes",
        "expected": "Near-zero delta across all sites under A486",
        "references": [
            "Lasko et al., Nature 2017 (PMID 29211713) — A486 inactive enantiomer",
        ],
    },
    "phospho_acetyl_crosstalk": {
        "description": "H3S10ph should anti-correlate with H3K9ac (known crosstalk)",
        "expected": "Cross-type attention between phospho and acetyl tokens",
        "references": [
            "Fischle et al., Nature 2003 (PMID 14573844) — H3 phospho-acetyl binary switch",
            "Ardito et al., IJMS 2019 — acetylation-phosphorylation crosstalk in HDAC inhibitors",
            "Zhang et al., Cell Death Discov 2025 — acetylation-autophagy crosstalk in cancer",
            "Nie et al., Cell Death Differ 2025 — O-GlcNAcylation-phosphorylation crosstalk",
        ],
    },
    "cudc101_dual_effect": {
        "description": "CUDC-101 targets both HDACs AND EGFR/HER2",
        "expected": "Acetylation increases + phosphorylation changes simultaneously",
        "references": [
            "Lai et al., J Med Chem 2010 (PMID 20568778) — CUDC-101 triple inhibitor",
            "Park et al., Nat Comm 2025 — Romidepsin renders cancer vulnerable to RTK targeting",
        ],
    },
    "hdac3_akt_crosstalk": {
        "description": "HDAC3-AKT interaction modulates chemoresistance",
        "expected": "AKT phospho sites co-vary with HDAC3 acetylation under HDAC inhibition",
        "references": [
            "Gupta et al., Leukemia 2017 — HDAC3-AKT partner in chemoresistance reversal",
        ],
    },
}

# ══════════════════════════════════════════════════════════════════════════════
# Literature References (Query 1: HDAC acetylation phosphorylation crosstalk)
# ══════════════════════════════════════════════════════════════════════════════

REFERENCES = {
    # ── Critical (data + methodology) ────────────────────────────────────────
    "hartl_2024": {
        "citation": "Hartl et al., Cell Reports 2024",
        "title": "Decrypting lysine deacetylase inhibitor action and protein modifications by dose-resolved proteomics",
        "url": "https://www.cell.com/cell-reports/fulltext/S2211-1247(24)00600-4",
        "relevance": "Dose-response HDAC inhibitor proteomics — potential multi-cell-line data for LOCLO",
        "use": "data_source, validation",
    },
    "liu_2025_jbc": {
        "citation": "Liu et al., JBC 2025",
        "title": "Integrating deep learning for PTM crosstalk on Hsp90 and drug binding",
        "url": "https://www.jbc.org/article/S0021-9258(25)02370-1/fulltext",
        "relevance": "DL method for PTM crosstalk + drug binding — direct competitor/related work",
        "use": "benchmark_comparison, architecture_rationale",
    },
    "ardito_2019": {
        "citation": "Ardito et al., IJMS 2019",
        "title": "The Crosstalk between Acetylation and Phosphorylation: Emerging New Roles for HDAC Inhibitors in the Heart",
        "url": "https://www.mdpi.com/1422-0067/20/1/102",
        "relevance": "Acetylation↔phosphorylation crosstalk under HDAC inhibitors — validates H3S10ph–K9ac switch",
        "use": "biology_validation",
    },
    # ── High (drug-specific validation) ──────────────────────────────────────
    "park_2025_natcomm": {
        "citation": "Park et al., Nat Comm 2025",
        "title": "The HDAC inhibitor romidepsin renders liver cancer vulnerable to RTK targeting and immunologically active",
        "url": "https://www.nature.com/articles/s41467-025-62934-0",
        "relevance": "Romidepsin + RTK sensitization — validates CUDC-101 dual mechanism",
        "use": "drug_validation",
    },
    # ── Moderate (background references) ─────────────────────────────────────
    "gupta_2017_leukemia": {
        "citation": "Gupta et al., Leukemia 2017",
        "title": "Targeting HDAC3, a new partner protein of AKT in the reversal of chemoresistance in AML",
        "url": "https://www.nature.com/articles/leu2017130",
        "relevance": "HDAC3-AKT crosstalk in chemoresistance — supports PTM crosstalk biology",
        "use": "biology_reference",
    },
    "zhang_2025_celldeath": {
        "citation": "Zhang et al., Cell Death Discovery 2025",
        "title": "Crosstalk between acetylation modification and autophagy in cancer",
        "url": "https://www.nature.com/articles/s41420-025-02809-x",
        "relevance": "Acetylation crosstalk review — supports multi-PTM modeling rationale",
        "use": "review_reference",
    },
    "wang_2025_hdac_review": {
        "citation": "Wang et al., Eur J Med Res 2025",
        "title": "Targeted intervention of tumor microenvironment with HDAC inhibitors and their combination therapy strategies",
        "url": "https://link.springer.com/article/10.1186/s40001-025-02326-8",
        "relevance": "HDAC inhibitor combination therapy review",
        "use": "review_reference",
    },
    "nie_2025_crosstalk": {
        "citation": "Nie et al., Cell Death Differ 2025",
        "title": "Crosstalk between O-GlcNAcylation and phosphorylation in metabolism: regulation and mechanism",
        "url": "https://www.nature.com/articles/s41418-025-01473-z",
        "relevance": "PTM crosstalk between glycosylation and phosphorylation — supports PTM-BDL multi-type architecture",
        "use": "architecture_rationale",
    },
    "wu_2024_biomarkers": {
        "citation": "Wu et al., Precision Clinical Medicine 2024",
        "title": "Protein modification systems as cancer biomarkers and therapeutic targets",
        "url": "https://academic.oup.com/pcm/article/9/2/pbag014/8666258",
        "relevance": "PTM systems as therapeutic targets — supports general thesis",
        "use": "review_reference",
    },
}

# ══════════════════════════════════════════════════════════════════════════════
# Hartl 2024 Validation Resource — Known Drug-Regulated PTM Sites
# ══════════════════════════════════════════════════════════════════════════════
# Source: Hartl et al., Cell Reports 2024 (mmc2 acetylation, mmc3 phospho)
# Cell line: Jurkat T cells (single cell line — validation only, not LOCLO)
# Use: Compare model IG attributions against independently identified sites
#      that are upregulated/downregulated by Vorinostat and Romidepsin.
# ══════════════════════════════════════════════════════════════════════════════

def load_hartl_validation(data_dir: str = "data/raw/drugptm/hartl_2024") -> dict:
    """
    Load Hartl 2024 known drug-regulated PTM sites for XAI validation.

    Returns dict with structure:
      {
        "acetylation": {
          "Vorinostat": {"upregulated": [...], "downregulated": [...], "total": int},
          "Romidepsin": {"upregulated": [...], "downregulated": [...], "total": int},
        },
        "phosphorylation": { ... same ... },
        "hdac_substrates": { ... from mmc4 PRM ... },
      }

    Each site entry has: gene_name, protein_site, sequence_window.
    """
    import os
    result = {"acetylation": {}, "phosphorylation": {}, "hdac_substrates": {}}

    try:
        import openpyxl
    except ImportError:
        return result

    # ── Acetylation (mmc2) ────────────────────────────────────────────────
    ac_path = os.path.join(data_dir, "mmc2.xlsx")
    if os.path.exists(ac_path):
        wb = openpyxl.load_workbook(ac_path, read_only=True)
        ws = wb["Acetylation sites"]
        for drug in ["Vorinostat", "Romidepsin"]:
            up_sites, down_sites, total = [], [], 0
            for row in ws.iter_rows(values_only=True, min_row=2):
                identified = str(row[9]) if row[9] else ""
                if drug not in identified:
                    continue
                total += 1
                gene = str(row[4]).split(";")[0].strip() if row[4] else ""
                site = str(row[0]) if row[0] else ""
                entry = {"gene": gene, "site": site}
                upregulated = str(row[11]) if row[11] else ""
                downregulated = str(row[13]) if row[13] else ""
                if drug in upregulated:
                    up_sites.append(entry)
                if drug in downregulated:
                    down_sites.append(entry)
            result["acetylation"][drug] = {
                "upregulated": up_sites, "downregulated": down_sites,
                "total": total,
                "n_up": len(up_sites), "n_down": len(down_sites),
            }
        wb.close()

    # ── Phosphorylation (mmc3) ────────────────────────────────────────────
    ph_path = os.path.join(data_dir, "mmc3.xlsx")
    if os.path.exists(ph_path):
        wb = openpyxl.load_workbook(ph_path, read_only=True)
        ws = wb["Phosphorylation sites"]
        for drug in ["Vorinostat", "Romidepsin"]:
            up_sites, down_sites, total = [], [], 0
            for row in ws.iter_rows(values_only=True, min_row=2):
                identified = str(row[9]) if row[9] else ""
                if drug not in identified:
                    continue
                total += 1
                gene = str(row[4]).split(";")[0].strip() if row[4] else ""
                site = str(row[0]) if row[0] else ""
                entry = {"gene": gene, "site": site}
                upregulated = str(row[11]) if row[11] else ""
                downregulated = str(row[13]) if row[13] else ""
                if drug in upregulated:
                    up_sites.append(entry)
                if drug in downregulated:
                    down_sites.append(entry)
            result["phosphorylation"][drug] = {
                "upregulated": up_sites, "downregulated": down_sites,
                "total": total,
                "n_up": len(up_sites), "n_down": len(down_sites),
            }
        wb.close()

    # ── HDAC substrate specificity (mmc4 PRM) ─────────────────────────────
    prm_path = os.path.join(data_dir, "mmc4.xlsx")
    if os.path.exists(prm_path):
        wb = openpyxl.load_workbook(prm_path, read_only=True)
        ws = wb["PRM results"]
        substrates = []
        for row in ws.iter_rows(values_only=True, min_row=2):
            if row[0]:
                substrates.append({
                    "uniprot": str(row[0]),
                    "gene": str(row[1]) if row[1] else "",
                    "peptide": str(row[2]) if row[2] else "",
                })
        result["hdac_substrates"] = {
            "HDAC1_HDAC6_targets": substrates,
            "n_targets": len(substrates),
        }
        wb.close()

    return result


# ══════════════════════════════════════════════════════════════════════════════
# External Benchmarking References (from Query 1 papers)
# ══════════════════════════════════════════════════════════════════════════════

BENCHMARK_METHODS = {
    "liu_2025_ptm_crosstalk_dl": {
        "citation": "Liu et al., JBC 2025",
        "title": "Integrating deep learning for PTM crosstalk on Hsp90 and drug binding",
        "url": "https://www.jbc.org/article/S0021-9258(25)02370-1/fulltext",
        "method_type": "DL + PTM crosstalk",
        "relevance": "Direct competitor — uses DL for PTM crosstalk + drug binding prediction",
        "comparison_notes": (
            "Focuses on Hsp90 PTM crosstalk, while PTM-BDL is protein-agnostic. "
            "PTM-BDL additionally uses typed self-attention for cross-PTM-type interactions "
            "and multi-modal fusion (sequence + structure + drug), which this method lacks."
        ),
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# PDB Structures
# ══════════════════════════════════════════════════════════════════════════════

HDAC_PDB_STRUCTURES = [
    {"id": "4LXZ", "description": "HDAC8 + Vorinostat (SAHA)", "drug": "Vorinostat"},
    {"id": "3MAX", "description": "HDAC7 + TSA (class II reference)", "drug": "TSA"},
    {"id": "5EDU", "description": "HDAC1 + Romidepsin-like peptide", "drug": "Romidepsin"},
    {"id": "4BKX", "description": "p300 HAT domain (A485 target)", "drug": "A485"},
]
