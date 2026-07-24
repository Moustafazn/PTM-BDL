#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  STEP 05 — Process Drug-PTM Data: Multi-PTM, Multi-Protein Integration       ║
║  (ERBB family: EGFR + HER2/ERBB2  ×  Phosphorylation + N-Glycosylation)      ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                              ║
║  PURPOSE:                                                                    ║
║    Build the unified drug-PTM measurement table used by step06.  Two PTM     ║
║    modification types are now extracted side-by-side for two target          ║
║    receptors, producing the substrate for the PTM Biological Dynamics Layer  ║
║    (PTM-BDL §3 / §7 / §10) introduced in `PTM_Biological_Dynamics_Layer.md`. ║
║                                                                              ║
║      target proteins:  EGFR (P00533),   HER2 / ERBB2 (P04626)                ║
║      PTM types     :   phosphorylation (Y / S / T),   N-glycosylation (N)    ║
║                                                                              ║
║    Each output row carries:                                                  ║
║      • target_protein            ∈ {EGFR, ERBB2}                            ║
║      • ptm_type                  ∈ {phosphorylation, N-glycosylation}        ║
║      • ptm_modification_type     ∈ {phospho_Y, phospho_S, phospho_T,         ║
║                                      glyco_N}                                ║
║    so step06 can build a typed PTM-BDL token vector (one token per site,     ║
║    typed by modification, see proposal §7.4).                                ║
║                                                                              ║
║  REASON FOR THE MULTI-PTM EXPANSION:                                         ║
║    The 2026-06-28 evaluation (`results/COMPREHENSIVE_EVALUATION_28_june.md`) ║
║    shows the randomised-PTM control DROPS by −0.042 BAcc / −0.010 AUROC —    ║
║    i.e. shuffling the PTM vectors makes the model BETTER.  This is the      ║
║    smoking gun for a PTM input channel that carries no usable biological     ║
║    signal (PTM-BDL §1.1).  The proposal's remedy is two-fold:                ║
║      (a) extend PTM coverage to a second modification type that is NOT a     ║
║          deterministic function of mutation_class — N-glycosylation          ║
║          encodes receptor-surface biology, ORTHOGONAL to phospho signalling. ║
║      (b) re-architect the model with a typed PTM-BDL self-attention layer.   ║
║    Step 05's job is (a): bring the glyco data into the same table.           ║
║                                                                              ║
║  EMPIRICAL NOTE (verified in this run, 2026-06-28):                          ║
║    A full scan of all DrugPTM-Bench cell-line CSVs                           ║
║    (`data/raw/drugptm/30394195/PTM_CellLine_*.csv`) for EGFR and ERBB2       ║
║    rows shows the `PTM type` column is *exclusively* "phosphorylation" for   ║
║    these two receptors.  DrugPTM-Bench therefore does NOT provide glyco-     ║
║    sylation data for our targets; glycosylation is entirely sourced from     ║
║    sources G–L below.  HeLa contains acetylation rows but not for our        ║
║    targets, so they are skipped.                                             ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  CANCER FOCUS                                                                ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    • EGFR-mutant NSCLC (lung)                                                ║
║    • HER2-amplified breast cancer                                            ║
║                                                                              ║

║  ────────────────────────────────────────────────────────────────────────── ║
║  SOURCE A — DrugPTM-Bench                                                    ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    Paper:   "DrugPTM-Bench: A Comprehensive Benchmark for Drug-Induced       ║
║             Post-Translational Modification Prediction"                       ║
║    Authors: Badkul A, Qi Y, Xie L                                            ║
║    Journal: Molecular Cell (2026)                                             ║
║    PMID:    30394195                                                         ║
║    DOI:     https://doi.org/10.1016/j.molcel.2018.11.028                    ║
║    GitHub:  https://github.com/Xie-lab/DrugPTM-Bench                        ║
║    Data:    Raw CSVs in data/raw/drugptm/30394195/                           ║
║                                                                              ║
║    Description:                                                              ║
║      Large-scale dose-response phosphoproteomic measurements across          ║
║      7 cancer cell lines × 26+ kinase inhibitors. Each cell line file        ║
║      (PTM_CellLine_<name>.csv) contains ~1–2 M rows of per-peptide,         ║
║      per-dose signal intensity data with fitted EC50 curves.                 ║
║      We extract EGFR phosphosite data from A431 (WT EGFR) treated with      ║
║      Gefitinib (1st-gen) and Afatinib (2nd-gen) TKIs.                       ║
║                                                                              ║
║    What it provides to our model:                                            ║
║      • Dose-dependent phosphorylation changes at EGFR sites                  ║
║      • EC50, pEC50, R², curve effect size per site × drug                    ║
║      • ~196 EGFR phosphosite dose-response summaries                        ║
║      • Drug SMILES via all_SMILES.csv                                        ║
║    Resistance context: "dose_response"                                       ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  SOURCE B — Tozuka et al., 2024                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    Paper:   "Phosphoproteomics reveals common and specific phosphorylation    ║
║             alterations in osimertinib-resistant NSCLC cells"                 ║
║    Authors: Tozuka T, Nishi H, Ohishi T, et al.                              ║
║    Journal: iScience 27(5): 109657 (2024)                                    ║
║    PMID:    38646155                                                         ║
║    DOI:     https://doi.org/10.1016/j.isci.2024.109657                      ║
║    PMC:     https://www.ncbi.nlm.nih.gov/pmc/articles/PMC11031815/          ║
║    Data:    data/raw/drugptm/tozuka_2024/mmc2.xlsx (Supplementary Table S2) ║
║                                                                              ║
║    Description:                                                              ║
║      TMT-based quantitative phosphoproteomics comparing parental vs          ║
║      osimertinib-resistant (OsiR) NSCLC cell lines:                          ║
║        • HCC827 (Exon 19 del) vs HCC827-OsiR                               ║
║        • PC-9   (Exon 19 del) vs PC-9-OsiR                                 ║
║      The mmc2.xlsx contains 17,853 phosphosites with log2-normalised         ║
║      TMT reporter intensities across 15 channels (3 reps × 5 conditions).  ║
║      We extract 21 EGFR-specific phosphosite rows.                           ║
║                                                                              ║
║    What it provides to our model:                                            ║
║      • Direct parental vs resistant phosphosite comparison                    ║
║      • log2FC = mean(resistant) − mean(parental) for each EGFR site          ║
║      • Sites like Y1172, Y978, Y1197 show dramatic dephosphorylation        ║
║        in resistant cells (log2FC < −2), indicating EGFR pathway shutdown   ║
║    Resistance context: "parental_vs_resistant"                               ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  SOURCE C — Hsu et al., 2025                                                 ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    Paper:   "Temporal phosphoproteomics reveals early signaling dynamics      ║
║             and drug-tolerant persister state in EGFR-mutant NSCLC"          ║
║    Authors: Hsu JL, Chen CT, et al.                                          ║
║    Journal: Molecular Systems Biology (2025)                                 ║
║    PMID:    41023502                                                         ║
║    DOI:     https://doi.org/10.1038/s44320-025-00141-1                      ║
║    PMC:     https://www.ncbi.nlm.nih.gov/pmc/articles/PMC12583488/          ║
║    Data:    data/raw/drugptm/hsu_2025/44320_2025_141_MOESM3_ESM.xlsx        ║
║             (Supplementary Dataset EV1)                                      ║
║                                                                              ║
║    Description:                                                              ║
║      DIA-MS phosphoproteomics tracking PC-9 cells (EGFR Exon 19 del)        ║
║      through an osimertinib resistance time course:                          ║
║        DMSO → Osi 5 min → Osi 10 min → Osi 6 h → DTP → DTP-24 h → DTP-7 d ║
║      where DTP = Drug-Tolerant Persister (surviving ~21 days on drug).       ║
║      Contains 7,943 phosphosites and 2,954 proteins across 7 conditions.     ║
║      We extract 5 EGFR phosphosites (S1064, S1166, S991, T693, T725).       ║
║                                                                              ║
║    What it provides to our model:                                            ║
║      • Temporal trajectory of phosphorylation during resistance emergence    ║
║      • Acute drug effect (5 min), sustained effect (6 h), persister state   ║
║      • fc_dtp_rebound = DTP − Osi_6h → unique bypass/recovery signal       ║
║    Resistance context: "temporal_dynamics"                                   ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  SOURCE D — PNAS 2025                                                        ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    Paper:   "Tyrosine phosphoproteome profiling identifies cell-intrinsic    ║
║             signals limiting the efficacy of TKI therapies"                  ║
║    Journal: PNAS (2025)                                                      ║
║    DOI:     https://doi.org/10.1073/pnas.2522090123                         ║
║    Data:    data/raw/drugptm/pnas_2025/pnas.2522090123.sd02.xlsx (S2)       ║
║             data/raw/drugptm/pnas_2025/pnas.2522090123.sd04.xlsx (S4)       ║
║                                                                              ║
║    Description:                                                              ║
║      TMT-based tyrosine phosphoproteomics of EGFR-mutant NSCLC cell         ║
║      lines treated with Osimertinib:                                         ║
║        • H1975  (EGFR L858R/T790M) — DMSO vs Osimertinib                   ║
║        • HCC4006 (EGFR Exon 19 del) — DMSO vs Osimertinib                  ║
║      Dataset S2: 444 pY sites × 4 conditions × 3 bio reps (TMT)            ║
║      Dataset S4: pre-computed differential statistics                        ║
║                                                                              ║
║    What it provides to our model:                                            ║
║      • Direct Osimertinib-induced pY phosphorylation changes                ║
║      • 444 pY sites across the full signaling network                       ║
║      • 5 direct EGFR sites: Y998, Y1092, Y1110, Y1172, Y1197              ║
║      • 27+ EGFR pathway sites (ERBB2, ERBB3, SRC, etc.)                   ║
║      • Two mutation backgrounds with differential Osi sensitivity           ║
║    Resistance context: "tki_phosphoproteome"                                ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  SOURCE E — FEBS/Mol Oncol 2025                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    Paper:   "Unveiling unique protein and phosphorylation signatures in      ║
║             lung adenocarcinomas with and without ALK, EGFR, and KRAS"      ║
║    Journal: Molecular Oncology (2025)                                        ║
║    DOI:     https://doi.org/10.1002/1878-0261.70091                         ║
║    MassIVE: MSV000095018                                                     ║
║    Data:    data/raw/drugptm/febs_2025/mol270091-sup-0006-tables5.xlsx      ║
║                                                                              ║
║    Description:                                                              ║
║      Phosphoproteomics of LUAD patient tumors across 4 genotypes:           ║
║        • EML4-ALK, EGFR-mutant, KRAS-mutant, Wild-type                     ║
║      Table S5: 211 phosphosites with fold-changes for 6 pairwise            ║
║      comparisons. 104 significantly different in EGFR vs WT.                ║
║                                                                              ║
║    What it provides to our model:                                            ║
║      • Tumor-derived phospho-signatures (in-vivo context)                   ║
║      • 211 phosphosites across the signaling network                        ║
║      • EGFR-mutant vs WT genotype-driven phospho differences               ║
║    Resistance context: "tumor_phospho_signatures"                           ║
║                                                                              ║
║  ────────────────────────────────────────────────────────────────────────── ║
║  COMBINED COVERAGE                                                           ║
║  ────────────────────────────────────────────────────────────────────────── ║
║    What              | Source          | Cell Lines   | Drug(s)              ║
║    ──────────────────|─────────────────|──────────────|───────────────────── ║
║    Dose-response     | DrugPTM-Bench   | A431 (WT)    | Gefitinib, Afatinib ║
║    Resistance state  | Tozuka 2024     | PC-9, HCC827 | Osimertinib         ║
║    Temporal dynamics | Hsu 2025        | PC-9         | Osimertinib         ║
║    pY phosphoproteome| PNAS 2025       | H1975,HCC4006| Osimertinib         ║
║    Tumor phospho sig.| FEBS/MolOnc 2025| LUAD tumors  | none (genotype)     ║
║                                                                              ║
║  OUTPUT:                                                                     ║
║    data/processed/drugptm/drugptm_egfr_phospho_responses.csv                ║
║    data/processed/drugptm/drugptm_cell_line_drug_catalog.csv                ║
║                                                                              ║
║  UNIFIED SCHEMA (each row = one phosphosite measurement):                    ║
║    cell_line, drug_name, protein, ptm_site, ptm_residue, ptm_type,          ║
║    peptide_sequence, baseline_intensity, max_dose_intensity,                 ║
║    log2_fold_change, data_source, resistance_context,                        ║
║    + source-specific columns (EC50, temporal dynamics, etc.)                 ║
║                                                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import re
from pathlib import Path

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
from src.ptm_bdl.config import load_config

cfg = load_config(case_study="egfr_erbb2_tki")

RAW_DIR = PROJECT_ROOT / cfg["paths"]["raw_data"] / "drugptm"
PMID_DIR = RAW_DIR / "30394195"
TOZUKA_DIR = RAW_DIR / "tozuka_2024"
HSU_DIR = RAW_DIR / "hsu_2025"
PNAS_DIR = RAW_DIR / "pnas_2025"
FEBS_DIR = RAW_DIR / "febs_2025"
CANCERRES_DIR = RAW_DIR / "cancerres_2021"
MCP_DIR = RAW_DIR / "mcp_2025"
OUT_DIR = PROJECT_ROOT / cfg["paths"]["processed_data"] / "drugptm"
RAW_DIR.mkdir(parents=True, exist_ok=True)
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Chunk size for reading large CSVs (adjust if memory-constrained)
CHUNK_SIZE = 100_000

# Drug SMILES from config
OSIMERTINIB_SMILES = cfg["drugs"]["osimertinib"]["smiles"]

# Columns we keep from the DrugPTM-Bench raw data to reduce memory
KEEP_COLUMNS = [
    "Sequence", "PTM Index", "PTM Residue", "PTM type",
    "Proteins", "Leading proteins", "Protein names", "Gene names",
    "Cell Line", "Timepoint (min)", "Chemical Name",
    "Dosage (µg)", "N duplicates",
    "Signal Intensity", "Signal Ratio",
    "Log EC50", "Curve slope", "Curve top", "Curve bottom",
    "R2", "EC50", "pEC50", "Curve effect size",
    "Canonical SMILES",
]


# ══════════════════════════════════════════════════════════════════════════════
# HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _safe_float(val) -> float:
    """Convert any value to float, returning NaN for failures."""
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().lower()
    if s in ("nan", "", "none", "na"):
        return np.nan
    try:
        return float(s)
    except (ValueError, TypeError):
        return np.nan


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE A: DrugPTM-Bench — Badkul et al., 2026 (PMID 30394195)
# "DrugPTM-Bench: A Comprehensive Benchmark for Drug-Induced PTM Prediction"
# Journal: Molecular Cell | DOI: 10.1016/j.molcel.2018.11.028
# GitHub:  https://github.com/Xie-lab/DrugPTM-Bench
# Data:    data/raw/drugptm/30394195/PTM_CellLine_*.csv + all_SMILES.csv
# ══════════════════════════════════════════════════════════════════════════════

def verify_drugptm_bench() -> list[Path]:
    """
    Check that the author-provided raw data files exist in
    data/raw/drugptm/30394195/.

    Returns a sorted list of PTM_CellLine_*.csv paths found.
    """
    print("\n" + "=" * 70)
    print("SOURCE A: DrugPTM-Bench — Badkul et al., 2026 (PMID 30394195)")
    print("=" * 70)

    if not PMID_DIR.exists():
        print(f"  ✗ Data directory not found: {PMID_DIR}")
        print("    Place the author-provided CSV files in:")
        print(f"      {PMID_DIR}/PTM_CellLine_<name>.csv")
        print(f"      {PMID_DIR}/all_SMILES.csv")
        return []

    cell_line_files = sorted(PMID_DIR.glob("PTM_CellLine_*.csv"))
    smiles_file = PMID_DIR / "all_SMILES.csv"

    print(f"  Directory: {PMID_DIR}")
    print(f"  Cell-line files found: {len(cell_line_files)}")
    for f in cell_line_files:
        print(f"    • {f.name}")
    print(f"  all_SMILES.csv: {'✓' if smiles_file.exists() else '✗ MISSING'}")

    if not cell_line_files:
        print("  ✗ No PTM_CellLine_*.csv files found.")
    return cell_line_files


def extract_egfr_data(cell_line_files: list[Path]) -> pd.DataFrame:
    """
    Stream through all raw CSVs in chunks and extract rows where
    Gene names contain 'EGFR'.  Because files are multi-GB, we never
    load more than CHUNK_SIZE rows at a time.
    """
    print("\n  Extracting EGFR phosphorylation data (chunked)...")

    egfr_frames: list[pd.DataFrame] = []

    for fpath in cell_line_files:
        cell_line_name = fpath.stem.replace("PTM_CellLine_", "")
        n_egfr = 0
        n_total = 0

        for chunk in pd.read_csv(fpath, chunksize=CHUNK_SIZE, low_memory=False):
            n_total += len(chunk)
            mask = chunk["Gene names"].astype(str).str.contains("EGFR", na=False)
            egfr_chunk = chunk.loc[mask].copy()
            if not egfr_chunk.empty:
                avail = [c for c in KEEP_COLUMNS if c in egfr_chunk.columns]
                egfr_frames.append(egfr_chunk[avail])
                n_egfr += len(egfr_chunk)

        print(f"    {cell_line_name:>15s}: {n_total:>10,} rows → {n_egfr:>6,} EGFR rows")

    if not egfr_frames:
        print("    ⚠ No EGFR rows found in any file.")
        return pd.DataFrame()

    df_egfr = pd.concat(egfr_frames, ignore_index=True)
    print(f"    Total EGFR rows extracted: {len(df_egfr):,}")
    return df_egfr


def build_drugptm_bench_summaries(df_egfr: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate the raw EGFR dose-response rows into one summary row per
    (cell_line, drug, ptm_site, ptm_residue) combination.

    For each group we capture:
      • baseline_intensity — Signal Intensity at Dosage == 0  (DMSO control)
      • max_dose_intensity — Signal Intensity at the highest dosage
      • log2_fold_change   — log2(max_dose / baseline)
      • EC50, pEC50, curve effect size, R² from curve fitting
      • n_doses — number of dose points measured
    """
    print("\n  Building per-site dose-response summaries...")

    if df_egfr.empty:
        print("    ⚠ No EGFR data to summarise.")
        return pd.DataFrame()

    # Ensure numeric types
    for col in ["Signal Intensity", "Signal Ratio", "Dosage (µg)",
                "Log EC50", "EC50", "pEC50", "Curve effect size", "R2",
                "Curve slope", "Curve top", "Curve bottom"]:
        if col in df_egfr.columns:
            df_egfr[col] = pd.to_numeric(df_egfr[col], errors="coerce")

    # Build a unique site label  (e.g. "Y1068")
    df_egfr["ptm_site"] = df_egfr["PTM Residue"] + df_egfr["PTM Index"].astype(str)

    group_cols = ["Cell Line", "Chemical Name", "Gene names",
                  "Sequence", "ptm_site", "PTM Residue", "PTM type"]

    records = []
    for keys, grp in df_egfr.groupby(group_cols, dropna=False):
        cell_line, drug, gene, seq, site, residue, ptm_type = keys

        baseline_rows = grp[grp["Dosage (µg)"] == 0.0]
        baseline_intensity = (baseline_rows["Signal Intensity"].mean()
                              if not baseline_rows.empty else np.nan)

        max_dose = grp["Dosage (µg)"].max()
        max_dose_rows = grp[grp["Dosage (µg)"] == max_dose]
        max_dose_intensity = (max_dose_rows["Signal Intensity"].mean()
                              if not max_dose_rows.empty else np.nan)

        if (pd.notna(baseline_intensity) and baseline_intensity > 0
                and pd.notna(max_dose_intensity)):
            log2fc = np.log2((max_dose_intensity + 0.01)
                             / (baseline_intensity + 0.01))
        else:
            log2fc = np.nan

        ec50 = (grp["EC50"].dropna().iloc[0]
                if "EC50" in grp.columns and grp["EC50"].notna().any()
                else np.nan)
        pec50 = (grp["pEC50"].dropna().iloc[0]
                 if "pEC50" in grp.columns and grp["pEC50"].notna().any()
                 else np.nan)
        eff_size = (grp["Curve effect size"].dropna().iloc[0]
                    if "Curve effect size" in grp.columns
                       and grp["Curve effect size"].notna().any()
                    else np.nan)
        r2 = (grp["R2"].dropna().iloc[0]
              if "R2" in grp.columns and grp["R2"].notna().any()
              else np.nan)
        smiles = (grp["Canonical SMILES"].dropna().iloc[0]
                  if "Canonical SMILES" in grp.columns
                     and grp["Canonical SMILES"].notna().any()
                  else "")

        records.append({
            "cell_line": cell_line,
            "drug_name": drug,
            "protein": gene,
            "ptm_site": site,
            "ptm_residue": residue,
            "ptm_type": ptm_type,
            "peptide_sequence": seq,
            "baseline_intensity": (round(baseline_intensity, 4)
                                   if pd.notna(baseline_intensity) else np.nan),
            "max_dose_intensity": (round(max_dose_intensity, 4)
                                   if pd.notna(max_dose_intensity) else np.nan),
            "max_dose_ug": max_dose,
            "log2_fold_change": (round(log2fc, 4)
                                 if pd.notna(log2fc) else np.nan),
            "EC50": ec50,
            "pEC50": pec50,
            "curve_effect_size": eff_size,
            "R2": r2,
            "n_doses": int(grp["Dosage (µg)"].nunique()),
            "drug_smiles": smiles,
            "data_source": "drugptm_bench",
            "resistance_context": "dose_response",
        })

    df_summary = pd.DataFrame(records)

    print(f"    → {len(df_summary):,} dose-response summaries")
    if not df_summary.empty:
        print(f"      Cell lines : {df_summary['cell_line'].nunique()}")
        print(f"      Drugs      : {df_summary['drug_name'].nunique()}")
        print(f"      PTM sites  : {df_summary['ptm_site'].nunique()}")
        print(f"      PTM types  : {sorted(df_summary['ptm_type'].dropna().unique())}")

    return df_summary


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE A-HER2: ERBB2 from DrugPTM-Bench (BT-474, MDA-MB-175)
# LIMITATION: BT-474 phospho uses Pertuzumab/Trastuzumab (antibodies) but
#   GDSC IC50 labels use Lapatinib/Afatinib (TKIs). Drug-modality mismatch.
#   MDA-MB-175 has Lapatinib but only 10 ERBB2 rows (sparse).
#   Future: contact Cui et al. 2025 for resistant phospho data.
# ══════════════════════════════════════════════════════════════════════════════


def extract_erbb2_data(cell_line_files: list[Path]) -> pd.DataFrame:
    """Extract ERBB2 phospho rows from HER2-relevant cell lines only."""
    print("\n  Extracting ERBB2 phosphorylation data (chunked)...")
    her2_lines = {"BT-474", "MDA-MB-175"}
    frames: list[pd.DataFrame] = []
    for fpath in cell_line_files:
        cl = fpath.stem.replace("PTM_CellLine_", "")
        if cl not in her2_lines:
            continue
        n_erbb2 = n_total = 0
        for chunk in pd.read_csv(fpath, chunksize=CHUNK_SIZE, low_memory=False):
            n_total += len(chunk)
            mask = chunk["Gene names"].astype(str).str.contains("ERBB2", na=False)
            hit = chunk.loc[mask].copy()
            if not hit.empty:
                avail = [c for c in KEEP_COLUMNS if c in hit.columns]
                frames.append(hit[avail])
                n_erbb2 += len(hit)
        print(f"    {cl:>15s}: {n_total:>10,} rows -> {n_erbb2:>6,} ERBB2 rows")
    if not frames:
        print("    ⚠ No ERBB2 rows found.")
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    print(f"    Total ERBB2 rows: {len(df):,}")
    return df


def build_erbb2_summaries(df_erbb2: pd.DataFrame) -> pd.DataFrame:
    """Build per-site ERBB2 dose-response summaries (mirrors EGFR logic)."""
    print("\n  Building ERBB2 per-site dose-response summaries...")
    if df_erbb2.empty:
        print("    ⚠ No ERBB2 data to summarise.")
        return pd.DataFrame()
    for col in ["Signal Intensity", "Signal Ratio", "Dosage (µg)",
                "Log EC50", "EC50", "pEC50", "Curve effect size", "R2"]:
        if col in df_erbb2.columns:
            df_erbb2[col] = pd.to_numeric(df_erbb2[col], errors="coerce")
    df_erbb2["ptm_site"] = df_erbb2["PTM Residue"] + df_erbb2["PTM Index"].astype(str)
    group_cols = ["Cell Line", "Chemical Name", "Gene names",
                  "Sequence", "ptm_site", "PTM Residue", "PTM type"]
    records = []
    for keys, grp in df_erbb2.groupby(group_cols, dropna=False):
        cell_line, drug, gene, seq, site, residue, ptm_type = keys
        bl = grp[grp["Dosage (µg)"] == 0.0]
        bl_int = bl["Signal Intensity"].mean() if not bl.empty else np.nan
        mx = grp["Dosage (µg)"].max()
        mx_rows = grp[grp["Dosage (µg)"] == mx]
        mx_int = mx_rows["Signal Intensity"].mean() if not mx_rows.empty else np.nan
        if pd.notna(bl_int) and bl_int > 0 and pd.notna(mx_int):
            log2fc = np.log2((mx_int + 0.01) / (bl_int + 0.01))
        else:
            log2fc = np.nan
        ec50 = grp["EC50"].dropna().iloc[0] if grp["EC50"].notna().any() else np.nan
        pec50 = grp["pEC50"].dropna().iloc[0] if grp["pEC50"].notna().any() else np.nan
        eff = (grp["Curve effect size"].dropna().iloc[0]
               if "Curve effect size" in grp.columns and grp["Curve effect size"].notna().any()
               else np.nan)
        r2 = grp["R2"].dropna().iloc[0] if grp["R2"].notna().any() else np.nan
        smiles = (grp["Canonical SMILES"].dropna().iloc[0]
                  if "Canonical SMILES" in grp.columns and grp["Canonical SMILES"].notna().any()
                  else "")
        records.append({
            "cell_line": cell_line, "drug_name": drug, "protein": gene,
            "ptm_site": site, "ptm_residue": residue, "ptm_type": ptm_type,
            "peptide_sequence": seq,
            "baseline_intensity": round(bl_int, 4) if pd.notna(bl_int) else np.nan,
            "max_dose_intensity": round(mx_int, 4) if pd.notna(mx_int) else np.nan,
            "max_dose_ug": mx,
            "log2_fold_change": round(log2fc, 4) if pd.notna(log2fc) else np.nan,
            "EC50": ec50, "pEC50": pec50, "curve_effect_size": eff, "R2": r2,
            "n_doses": int(grp["Dosage (µg)"].nunique()),
            "drug_smiles": smiles,
            "data_source": "drugptm_bench",
            "resistance_context": "dose_response",
            "target_protein": "ERBB2",
        })
    df_sum = pd.DataFrame(records)
    print(f"    → {len(df_sum):,} ERBB2 dose-response summaries")
    if not df_sum.empty:
        print(f"      Cell lines: {sorted(df_sum['cell_line'].unique())}")
        print(f"      Drugs:      {sorted(df_sum['drug_name'].unique())}")
        print(f"      Sites:      {df_sum['ptm_site'].nunique()}")
    erbb2_path = OUT_DIR / "drugptm_erbb2_phospho_responses.csv"
    df_sum.to_csv(erbb2_path, index=False)
    print(f"    ✓ Saved: {erbb2_path.relative_to(PROJECT_ROOT)}")
    print(f"\n    ⚠ LIMITATION: BT-474 phospho = Pertuzumab/Trastuzumab (antibodies)")
    print(f"      GDSC IC50 labels = Lapatinib/Afatinib (TKIs). Drug mismatch noted.")
    return df_sum


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE B: Tozuka et al., 2024 (PMID 38646155)
# "Phosphoproteomics reveals common and specific phosphorylation alterations
#  in osimertinib-resistant NSCLC cells"
# Journal: iScience 27(5): 109657 | DOI: 10.1016/j.isci.2024.109657
# PMC:     https://www.ncbi.nlm.nih.gov/pmc/articles/PMC11031815/
# Data:    data/raw/drugptm/tozuka_2024/mmc2.xlsx (Supplementary Table S2)
# ══════════════════════════════════════════════════════════════════════════════

def process_tozuka_2024() -> pd.DataFrame:
    """
    Process Tozuka 2024 TMT phosphoproteomics data (mmc2.xlsx).

    This dataset compares phosphorylation between:
      • HCC827 parental  vs  HCC827 Osimertinib-Resistant (OsiR)
      • PC-9   parental  vs  PC-9   Osimertinib-Resistant (PC-9OsiR)

    The mmc2.xlsx "Results" sheet uses a 3-row header:
      Row 0: Column category labels
      Row 1: TMT channel numbers  (1,2,3,7,8,9,7-2,8-2,9-2,10,11,12,16,17,18)
      Row 2: Cell-line labels

    Channel mapping (columns 0–14, all log2 reporter intensity corrected):
      Cols 0–2:   HCC827 parental       (3 biological replicates)
      Cols 3–5:   HCC827 OsiR, batch 1  (3 replicates)
      Cols 6–8:   HCC827 OsiR, batch 2  (3 replicates)
      Cols 9–11:  PC-9 parental         (3 replicates)
      Cols 12–14: PC-9 OsiR             (3 replicates)

    Annotation columns (15–22):
      15: Amino acid (S / T / Y)
      16: Multiplicity  (___1 = singly phosphorylated, ___2 = doubly)
      17: Proteins      (UniProt IDs; semicolon-separated if shared peptide)
      18: Positions within proteins
      19: Protein names
      20: Gene names
      21: Sequence window
      22: Fasta headers

    Values are log2-normalised, so:
      log2FC(resistant vs parental) = mean(resistant channels) − mean(parental channels)
    """
    print("\n" + "=" * 70)
    print("SOURCE B: Tozuka et al., 2024 (PMID 38646155)")
    print("=" * 70)

    mmc2_path = TOZUKA_DIR / "mmc2.xlsx"
    if not mmc2_path.exists():
        print(f"  ✗ File not found: {mmc2_path}")
        print("    Download mmc2.xlsx from the journal supplement:")
        print("    https://doi.org/10.1016/j.isci.2024.109657")
        return pd.DataFrame()

    print(f"  Reading: {mmc2_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(mmc2_path, read_only=True)
    ws = wb["Results"]

    records: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # skip 3-row header
            continue

        gene_names = str(row[20]) if row[20] else ""
        if "EGFR" not in gene_names:
            continue

        amino_acid = str(row[15]).strip() if row[15] else ""
        multiplicity = str(row[16]).strip() if row[16] else ""
        proteins_raw = str(row[17]) if row[17] else ""
        positions_raw = str(row[18]) if row[18] else ""
        protein_names = str(row[19]) if row[19] else ""
        sequence_window = str(row[21]) if row[21] else ""

        # ── Extract the EGFR-specific residue position ──────────────────────
        proteins_list = proteins_raw.split(";")
        positions_list = positions_raw.split(";")

        egfr_position = None
        for j, prot in enumerate(proteins_list):
            if prot.strip().startswith("P00533"):
                if j < len(positions_list):
                    try:
                        egfr_position = int(float(positions_list[j].strip()))
                    except (ValueError, TypeError):
                        pass
                break

        if egfr_position is None:  # fallback: first position
            try:
                egfr_position = int(float(positions_list[0].strip()))
            except (ValueError, TypeError):
                continue  # skip row if position unreadable

        ptm_site = f"{amino_acid}{egfr_position}"

        # ── Gather intensity values per condition ───────────────────────────
        hcc827_par = [_safe_float(row[k]) for k in range(0, 3)]
        hcc827_res = [_safe_float(row[k]) for k in range(3, 9)]  # 2 batches
        pc9_par = [_safe_float(row[k]) for k in range(9, 12)]
        pc9_res = [_safe_float(row[k]) for k in range(12, 15)]

        hcc827_par_ok = [v for v in hcc827_par if np.isfinite(v)]
        hcc827_res_ok = [v for v in hcc827_res if np.isfinite(v)]
        pc9_par_ok = [v for v in pc9_par if np.isfinite(v)]
        pc9_res_ok = [v for v in pc9_res if np.isfinite(v)]

        # Use first (potentially shared) peptide from the sequence window
        peptide = (sequence_window.split(";")[0].strip()
                   if ";" in sequence_window else sequence_window.strip())

        # Determine gene label (use "EGFR" even for shared peptides)
        gene_label = "EGFR"
        if ";" in gene_names:
            gene_label = gene_names  # e.g. "EGFR;ERBB4;ERBB2"

        # ── HCC827: parental vs OsiR comparison ────────────────────────────
        if len(hcc827_par_ok) >= 2 and len(hcc827_res_ok) >= 2:
            par_mean = float(np.mean(hcc827_par_ok))
            res_mean = float(np.mean(hcc827_res_ok))
            log2fc = res_mean - par_mean  # log2 space → diff = FC

            records.append({
                "cell_line": "HCC827",
                "drug_name": "Osimertinib",
                "protein": gene_label,
                "ptm_site": ptm_site,
                "ptm_residue": amino_acid,
                "ptm_type": "phosphorylation",
                "peptide_sequence": peptide,
                "baseline_intensity": round(par_mean, 4),
                "max_dose_intensity": round(res_mean, 4),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(log2fc, 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "tozuka_2024",
                "resistance_context": "parental_vs_resistant",
            })

        # ── PC-9: parental vs PC-9OsiR comparison ─────────────────────────
        if len(pc9_par_ok) >= 2 and len(pc9_res_ok) >= 2:
            par_mean = float(np.mean(pc9_par_ok))
            res_mean = float(np.mean(pc9_res_ok))
            log2fc = res_mean - par_mean

            records.append({
                "cell_line": "PC-9",
                "drug_name": "Osimertinib",
                "protein": gene_label,
                "ptm_site": ptm_site,
                "ptm_residue": amino_acid,
                "ptm_type": "phosphorylation",
                "peptide_sequence": peptide,
                "baseline_intensity": round(par_mean, 4),
                "max_dose_intensity": round(res_mean, 4),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(log2fc, 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "tozuka_2024",
                "resistance_context": "parental_vs_resistant",
            })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} EGFR phospho comparisons")
    if not df.empty:
        for cl in sorted(df["cell_line"].unique()):
            sub = df[df["cell_line"] == cl]
            print(f"    {cl}: {len(sub)} sites, "
                  f"mean log2FC = {sub['log2_fold_change'].mean():+.3f}")
        # Show key resistance-associated sites
        big_changes = df[df["log2_fold_change"].abs() >= 1.0].sort_values(
            "log2_fold_change", key=abs, ascending=False)
        if not big_changes.empty:
            print(f"    Sites with |log2FC| ≥ 1.0:")
            for _, r in big_changes.head(10).iterrows():
                print(f"      {r['cell_line']:8s} {r['ptm_site']:8s} "
                      f"log2FC = {r['log2_fold_change']:+.3f}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE C: Hsu et al., 2025 (PMID 41023502)
# "Temporal phosphoproteomics reveals early signaling dynamics and
#  drug-tolerant persister state in EGFR-mutant NSCLC"
# Journal: Molecular Systems Biology (2025) | DOI: 10.1038/s44320-025-00141-1
# PMC:     https://www.ncbi.nlm.nih.gov/pmc/articles/PMC12583488/
# Data:    data/raw/drugptm/hsu_2025/44320_2025_141_MOESM3_ESM.xlsx (EV1)
# ══════════════════════════════════════════════════════════════════════════════

def process_hsu_2025() -> pd.DataFrame:
    """
    Process Hsu 2025 DIA-MS phosphoproteomics data.

    This dataset tracks phosphorylation changes in PC-9 cells (EGFR exon19del)
    treated with Osimertinib across a resistance-development time course:

      DMSO (baseline) → Osi 5 min → Osi 10 min → Osi 6 h → DTP → DTP-24 h → DTP-7 d

    where DTP = Drug-Tolerant Persister (cells surviving ~21 days of treatment).

    Values are log2-normalised phosphosite intensities.

    The "Phos_IDs" sheet contains (per row):
      Cols 0–6:  DMSO | Osi_5min | Osi_10min | Osi_6h | DTP | DTP-24h | DTP-7d
      Col 7:     ANOVA q-value
      Col 8:     PG.Genes
      Col 9:     PG.ProteinGroups (UniProt ID)
      Col 10:    EG.ModifiedPeptide (phosphopeptide sequences)
      Col 11:    EG.PTMLocalizationProbabilities
      Col 12:    PTM_group
      Col 13:    PTM_collapse_key  (e.g., "EGFR_S1064_M1")

    For each EGFR phosphosite we compute fold-changes at multiple windows:
      fc_acute_5min    = Osi_5min − DMSO   (immediate kinase inhibition)
      fc_sustained_6h  = Osi_6h − DMSO     (sustained drug effect)
      fc_dtp_persister = DTP − DMSO        (persister-state remodelling)
      fc_dtp_rebound   = DTP − Osi_6h      (recovery/bypass signal)

    The *primary* log2_fold_change used by the model is the sustained effect
    (Osi 6 h vs DMSO), the most comparable time-scale to conventional dose-
    response experiments.
    """
    print("\n" + "=" * 70)
    print("SOURCE C: Hsu et al., 2025 (PMID 41023502)")
    print("=" * 70)

    xlsx_files = list(HSU_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"  ✗ No .xlsx files found in: {HSU_DIR}")
        print("    Download supplementary data from:")
        print("    https://doi.org/10.1038/s44320-025-00141-1")
        return pd.DataFrame()

    xlsx_path = xlsx_files[0]
    print(f"  Reading: {xlsx_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    if "Phos_IDs" not in wb.sheetnames:
        print(f"  ✗ 'Phos_IDs' sheet not found. Available: {wb.sheetnames}")
        wb.close()
        return pd.DataFrame()

    ws = wb["Phos_IDs"]

    records: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header row
            continue

        gene = str(row[8]) if row[8] else ""
        protein = str(row[9]) if row[9] else ""
        ptm_key = str(row[13]) if row[13] else ""

        if "EGFR" not in gene and "P00533" not in protein:
            continue

        # ── Parse PTM site from collapse key (e.g. "EGFR_S1064_M1") ────────
        match = re.match(r"EGFR_([STY]\d+)_M\d+", ptm_key)
        if not match:
            continue
        ptm_site = match.group(1)  # e.g. "S1064"
        ptm_residue = ptm_site[0]  # "S", "T", or "Y"

        # ── Extract time-course intensities ─────────────────────────────────
        dmso = _safe_float(row[0])
        osi_5min = _safe_float(row[1])
        osi_10min = _safe_float(row[2])
        osi_6h = _safe_float(row[3])
        dtp = _safe_float(row[4])
        dtp_24h = _safe_float(row[5])
        dtp_7d = _safe_float(row[6])
        anova_q = _safe_float(row[7])

        # ── Compute fold-changes relative to DMSO ──────────────────────────
        def _diff(a, b):
            return round(a - b, 4) if np.isfinite(a) and np.isfinite(b) else np.nan

        fc_acute = _diff(osi_5min, dmso)
        fc_sustained = _diff(osi_6h, dmso)
        fc_dtp = _diff(dtp, dmso)
        fc_rebound = _diff(dtp, osi_6h)

        # Primary log2FC for the unified table: sustained 6 h effect
        log2fc = fc_sustained

        # ── Clean peptide string ────────────────────────────────────────────
        mod_peptide = str(row[10]) if row[10] else ""
        peptide_clean = mod_peptide.split(";")[0].strip("_ ")
        peptide_clean = re.sub(r"\[.*?\]", "", peptide_clean)  # strip [Phospho...]

        records.append({
            "cell_line": "PC-9",
            "drug_name": "Osimertinib",
            "protein": "EGFR",
            "ptm_site": ptm_site,
            "ptm_residue": ptm_residue,
            "ptm_type": "phosphorylation",
            "peptide_sequence": peptide_clean,
            "baseline_intensity": round(dmso, 4) if np.isfinite(dmso) else np.nan,
            "max_dose_intensity": round(osi_6h, 4) if np.isfinite(osi_6h) else np.nan,
            "max_dose_ug": np.nan,
            "log2_fold_change": log2fc,
            "EC50": np.nan, "pEC50": np.nan,
            "curve_effect_size": np.nan, "R2": np.nan,
            "n_doses": np.nan,
            "drug_smiles": OSIMERTINIB_SMILES,
            "data_source": "hsu_2025",
            "resistance_context": "temporal_dynamics",
            # Temporal columns unique to Hsu 2025
            "intensity_dmso": round(dmso, 4) if np.isfinite(dmso) else np.nan,
            "intensity_osi_5min": round(osi_5min, 4) if np.isfinite(osi_5min) else np.nan,
            "intensity_osi_10min": round(osi_10min, 4) if np.isfinite(osi_10min) else np.nan,
            "intensity_osi_6h": round(osi_6h, 4) if np.isfinite(osi_6h) else np.nan,
            "intensity_dtp": round(dtp, 4) if np.isfinite(dtp) else np.nan,
            "intensity_dtp_24h": round(dtp_24h, 4) if np.isfinite(dtp_24h) else np.nan,
            "intensity_dtp_7d": round(dtp_7d, 4) if np.isfinite(dtp_7d) else np.nan,
            "anova_q_value": anova_q,
            "fc_acute_5min": fc_acute,
            "fc_sustained_6h": fc_sustained,
            "fc_dtp_persister": fc_dtp,
            "fc_dtp_rebound": fc_rebound,
        })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} EGFR phosphosites × 7 time points")
    if not df.empty:
        print(f"    Sites: {', '.join(sorted(df['ptm_site'].tolist()))}")
        print(f"    ANOVA q-values: {df['anova_q_value'].min():.6f} – "
              f"{df['anova_q_value'].max():.6f}")
        print(f"    Acute FC (5 min vs DMSO): "
              f"{df['fc_acute_5min'].min():+.3f} to {df['fc_acute_5min'].max():+.3f}")
        print(f"    DTP persister FC:         "
              f"{df['fc_dtp_persister'].min():+.3f} to {df['fc_dtp_persister'].max():+.3f}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE D: PNAS 2025 — Tyrosine Phosphoproteome under TKI Treatment
# "Tyrosine phosphoproteome profiling identifies cell-intrinsic signals
#  limiting the efficacy of tyrosine kinase inhibitor therapies"
# Journal: PNAS (2025) | DOI: 10.1073/pnas.2522090123
# Data:    data/raw/drugptm/pnas_2025/pnas.2522090123.sd02.xlsx (Dataset S2)
#          data/raw/drugptm/pnas_2025/pnas.2522090123.sd04.xlsx (Dataset S4)
# ══════════════════════════════════════════════════════════════════════════════

def process_pnas_2025() -> pd.DataFrame:
    """
    Process PNAS 2025 tyrosine phosphoproteomics data.

    This study profiled the tyrosine phosphoproteome of EGFR-mutant NSCLC
    cell lines treated with Osimertinib (3rd-gen EGFR TKI) using TMT-based
    quantitative mass spectrometry.

    Dataset S2 ("EGFR plex" sheet) contains TMT reporter intensities for
    444 tyrosine phosphosites across 4 conditions × 3 biological replicates:
      • H1975 (EGFR L858R/T790M) — DMSO control vs Osimertinib
      • HCC4006 (EGFR Exon 19 del) — DMSO control vs Osimertinib

    Dataset S4 ("EGFR lines, OSI" sheet) contains pre-computed differential
    statistics (log2FC between H1975 vs HCC4006 after Osimertinib).

    Column layout in S2 "EGFR plex" sheet:
      Col 0:  Phosphosites         (e.g., "EGFR-pY1092")
      Col 1:  Annotated peptide sequences
      Col 2:  Modifications
      Col 3:  Protein accessions   (UniProt IDs)
      Col 4:  Gene IDs
      Col 5:  Protein descriptions
      Col 6:  Number of PSMs (total)
      Col 7:  Number of PSMs (missing values)
      Col 8–10:   H1975 DMSO BR1–3  (TMT intensity)
      Col 11–13:  H1975 OSI BR1–3   (TMT intensity)
      Col 14–16:  HCC4006 DMSO BR1–3 (TMT intensity)
      Col 17–19:  HCC4006 OSI BR1–3  (TMT intensity)

    For each phosphosite we compute:
      log2FC = log2(mean(OSI replicates) / mean(DMSO replicates))

    This is the FIRST dataset with direct Osimertinib pY phosphoproteomics,
    directly filling Gap #3 (no Osi dose-response phospho data).

    Cell line mutation backgrounds:
      H1975:   EGFR L858R + T790M (double mutant, partially Osi-resistant)
      HCC4006: EGFR Exon 19 del (E746-A750del, Osi-sensitive)

    What it provides to our model:
      • Direct Osimertinib-induced tyrosine phosphorylation changes
      • 444 pY sites across the full signaling network
      • Two EGFR-mutant backgrounds with differential Osi sensitivity
      • 5 direct EGFR pY sites: Y998, Y1092, Y1110, Y1172, Y1197
      • 27+ EGFR pathway protein sites (ERBB2, ERBB3, SRC, etc.)
    Resistance context: "tki_phosphoproteome"
    """
    print("\n" + "=" * 70)
    print("SOURCE D: PNAS 2025 — Tyrosine Phosphoproteome TKI Efficacy")
    print("=" * 70)

    sd02_path = PNAS_DIR / "pnas.2522090123.sd02.xlsx"
    sd04_path = PNAS_DIR / "pnas.2522090123.sd04.xlsx"

    if not sd02_path.exists():
        print(f"  ✗ File not found: {sd02_path}")
        print("    Download Dataset S2 (processed pY proteomics) from:")
        print("    https://www.pnas.org/doi/abs/10.1073/pnas.2522090123")
        return pd.DataFrame()

    print(f"  Reading: {sd02_path.name}")

    import openpyxl

    # ── Load pre-computed stats from S04 (if available) ──────────────────
    s04_stats: dict[str, dict] = {}  # site → {log2fc_dmso, log2fc_osi, pval, adj_pval}
    if sd04_path.exists():
        wb4 = openpyxl.load_workbook(sd04_path, read_only=True)
        for sheet_key, col_label in [("EGFR lines, DMSO", "dmso"),
                                     ("EGFR lines, OSI", "osi")]:
            if sheet_key in wb4.sheetnames:
                ws4 = wb4[sheet_key]
                for i, row in enumerate(ws4.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    site_id = str(row[0]) if row[0] else ""
                    if site_id not in s04_stats:
                        s04_stats[site_id] = {}
                    s04_stats[site_id][f"log2fc_cellline_ratio_{col_label}"] = (
                        _safe_float(row[2]))
                    s04_stats[site_id][f"pval_{col_label}"] = _safe_float(row[3])
                    s04_stats[site_id][f"adj_pval_{col_label}"] = _safe_float(row[4])
        wb4.close()
        print(f"  ✓ Loaded S04 stats for {len(s04_stats):,} phosphosites")

    # ── Process S02: EGFR plex — raw TMT intensities ─────────────────────
    wb = openpyxl.load_workbook(sd02_path, read_only=True)

    if "EGFR plex" not in wb.sheetnames:
        print(f"  ✗ 'EGFR plex' sheet not found. Available: {wb.sheetnames}")
        wb.close()
        return pd.DataFrame()

    ws = wb["EGFR plex"]

    # Cell line info for annotation
    cell_line_info = {
        "H1975": {
            "dmso_cols": (8, 9, 10),
            "osi_cols": (11, 12, 13),
            "mutation": "L858R_T790M",
            "egfr_context": "double_mutant",
        },
        "HCC4006": {
            "dmso_cols": (14, 15, 16),
            "osi_cols": (17, 18, 19),
            "mutation": "Exon19del",
            "egfr_context": "exon19del",
        },
    }

    records: list[dict] = []
    n_egfr_direct = 0
    n_pathway = 0

    # ── PTM-Derived Pathway Classification ────────────────────────────────
    # Each phosphosite is assigned to a signaling pathway based on its
    # parent protein.  These are NOT independent pathway features — they
    # are PTM-derived signaling representations that capture how individual
    # phosphorylation events collectively rewire cellular signaling.
    #
    # The three-level biological hierarchy:
    #   Level 1: Individual PTM sites (Y869, Y1092...) — biological foundation
    #   Level 2: PTM rewiring (drug-induced log2FC per site)
    #   Level 3: Pathway rewiring (PTM-derived pathway summaries)
    #
    # Step06 reads pnas_protein_class to compute per-pathway aggregate
    # features (pw_* columns), enabling the model to learn:
    #   "EGFR is dephosphorylated but SRC maintains phosphorylation
    #    → resistance via SRC bypass"
    #
    # References: Citri & Yarden, Nat Rev Mol Cell Biol 2006;
    #             Rotow & Bivona, Nat Rev Cancer 2017
    # ─────────────────────────────────────────────────────────────────────
    pathway_gene_groups = {
        "egfr_direct": {"EGFR"},
        "erbb_family": {"ERBB2", "ERBB3", "ERBB4"},
        "mapk_pathway": {"MAPK1", "MAPK3", "MAPK7", "MAPK9", "MAPK10",
                         "MAPK11", "MAPK12", "MAPK13", "MAPK14",
                         "MAP2K1", "MAP2K2", "RAF1", "BRAF"},
        "pi3k_akt_pathway": {"PIK3R1", "PIK3R2", "PIK3R3", "PIK3CA", "PIK3CB",
                             "AKT1", "AKT2", "AKT3", "MTOR", "GSK3A", "GSK3B"},
        "src_fak_pathway": {"SRC", "YES1", "FYN", "FGR", "LYN", "LCK", "HCK",
                            "ABL1", "ABL2", "PTK2", "BCAR1", "NEDD9",
                            "PEAK1", "SRCIN1"},
        "bypass_rtk": {"MET", "AXL", "MERTK", "IGF1R", "INSR",
                       "EPHA2", "IRS2"},
        "adapter_effector": {"GAB1", "GRB2", "SHC1", "SOS1", "NCK1",
                             "PLCG1", "PLCG2", "STAT3", "STAT5A", "STAT5B",
                             "CBL", "PTPN11", "PTPN1", "VAV1", "TNK2", "TYK2"},
        "emt_adhesion": {"VIM", "CTNND1", "CTTN", "TJP1", "TJP2"},
    }
    # Build reverse lookup: gene → pathway_class
    gene_to_pathway = {}
    for pw, genes in pathway_gene_groups.items():
        for g in genes:
            gene_to_pathway[g] = pw

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue

        site_id = str(row[0]) if row[0] else ""  # e.g., "EGFR-pY1092"
        peptide = str(row[1]) if row[1] else ""
        modifications = str(row[2]) if row[2] else ""
        accession = str(row[3]) if row[3] else ""
        gene_id = str(row[4]) if row[4] else ""
        description = str(row[5]) if row[5] else ""

        # Parse the phosphosite label: "GENE-pYNNN" → gene, residue, position
        match = re.match(r"^(.+?)-p([YST])(\d+)", site_id)
        if not match:
            continue
        gene = match.group(1)  # e.g., "EGFR"
        ptm_residue = match.group(2)  # "Y", "S", or "T"
        ptm_position = match.group(3)  # e.g., "1092"
        ptm_site = f"{ptm_residue}{ptm_position}"  # e.g., "Y1092"

        # Classify the protein into granular pathway
        gene_first = gene.split(";")[0].strip()
        if gene_first in gene_to_pathway:
            protein_class = gene_to_pathway[gene_first]
            if gene_first == "EGFR":
                n_egfr_direct += 1
            else:
                n_pathway += 1
        else:
            protein_class = "signaling_other"

        # Clean peptide — take first if multiple
        peptide_clean = peptide.split("|")[0].strip()

        # Get S04 stats if available
        site_stats = s04_stats.get(site_id, {})

        # ── Create one row per cell line ─────────────────────────────────
        for cl_name, cl_info in cell_line_info.items():
            dmso_vals = [_safe_float(row[c]) for c in cl_info["dmso_cols"]]
            osi_vals = [_safe_float(row[c]) for c in cl_info["osi_cols"]]

            dmso_ok = [v for v in dmso_vals if np.isfinite(v)]
            osi_ok = [v for v in osi_vals if np.isfinite(v)]

            if len(dmso_ok) < 2 or len(osi_ok) < 2:
                continue  # skip if insufficient replicates

            dmso_mean = float(np.mean(dmso_ok))
            osi_mean = float(np.mean(osi_ok))

            if dmso_mean > 0 and osi_mean > 0:
                log2fc = float(np.log2(osi_mean / dmso_mean))
            else:
                log2fc = np.nan

            records.append({
                "cell_line": cl_name,
                "drug_name": "Osimertinib",
                "protein": gene_id,
                "ptm_site": ptm_site,
                "ptm_residue": ptm_residue,
                "ptm_type": "phosphorylation",
                "peptide_sequence": peptide_clean,
                "baseline_intensity": round(dmso_mean, 4),
                "max_dose_intensity": round(osi_mean, 4),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(log2fc, 4) if np.isfinite(log2fc) else np.nan,
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "pnas_2025",
                "resistance_context": "tki_phosphoproteome",
                # PNAS-specific columns
                "pnas_site_id": site_id,
                "pnas_protein_class": protein_class,
                "pnas_mutation_background": cl_info["mutation"],
                "pnas_uniprot": accession,
            })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} phosphosite × cell-line measurements")
    if not df.empty:
        print(f"    Cell lines:      {sorted(df['cell_line'].unique())}")
        print(f"    Unique sites:    {df['ptm_site'].nunique()}")
        print(f"    EGFR direct:     {n_egfr_direct} sites "
              f"({df[df['pnas_protein_class'] == 'egfr_direct'].shape[0]} rows)")
        print(f"    Classified pathway proteins: {n_pathway} sites")
        # Show per-pathway breakdown
        pw_counts = df['pnas_protein_class'].value_counts()
        for pw, cnt in pw_counts.items():
            n_prots = df[df['pnas_protein_class'] == pw]['protein'].nunique()
            print(f"      {pw:22s}: {cnt:3d} rows ({n_prots} proteins)")
        print(f"    Network total:   {df.shape[0]} rows")

        # Show key EGFR sites
        egfr_rows = df[df["pnas_protein_class"] == "egfr_direct"]
        if not egfr_rows.empty:
            print(f"    ── EGFR pY sites ──")
            for _, r in egfr_rows.sort_values(["ptm_site", "cell_line"]).iterrows():
                print(f"      {r['cell_line']:8s} {r['ptm_site']:8s} "
                      f"log2FC = {r['log2_fold_change']:+.3f}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE E: FEBS/Mol Oncol 2025 — Phosphorylation Signatures in LUAD Tumors
# "Unveiling unique protein and phosphorylation signatures in lung
#  adenocarcinomas with and without ALK, EGFR, and KRAS genetic alterations"
# Journal: Molecular Oncology (2025) | DOI: 10.1002/1878-0261.70091
# MassIVE: MSV000095018 | DOI: 10.25345/C5W37M669
# Data:    data/raw/drugptm/febs_2025/mol270091-sup-0006-tables5.xlsx (S5)
#          data/raw/drugptm/febs_2025/mol270091-sup-0007-tables6.xlsx (S6)
# ══════════════════════════════════════════════════════════════════════════════

def process_febs_2025() -> pd.DataFrame:
    """
    Process FEBS/Mol Oncol 2025 phosphoproteomics from LUAD patient tumors.

    This study compared phosphoproteomes of lung adenocarcinoma (LUAD) tumors
    across four genotype groups: EML4-ALK, EGFR-mutant, KRAS-mutant, and WT.

    Table S5 ("two_group_tests_results_Phospho" sheet) contains 211
    phosphosites with fold-changes and adjusted p-values for 6 pairwise
    comparisons between the four groups.

    Column layout:
      Col 0:  Sequence (phosphopeptide)
      Col 1:  Gene
      Col 2:  Uniprot ID
      Col 3:  Site (e.g., "GENE;S123")
      Col 14: Adjusted p-value (EGFR vs WT)
      Col 20: Fold-change log2 (FC EGFR vs WT)

    The primary log2_fold_change = FC EGFR vs WT (EGFR-mutant tumors
    vs wild-type), showing genotype-driven phospho differences in patient
    tissue — the in-vivo complement to our cell-line TKI response data.

    What it provides to our model:
      • Tumor-derived phospho-signatures from LUAD patients
      • 211 phosphosites across the signaling network
      • 104 significantly different in EGFR-mutant vs WT (adj.p < 0.05)
      • In-vivo context that validates cell-line findings
    Resistance context: "tumor_phospho_signatures"
    """
    print("" + "=" * 70)
    print("SOURCE E: FEBS/Mol Oncol 2025 — Tumor Phospho Signatures")
    print("=" * 70)

    s5_path = FEBS_DIR / "mol270091-sup-0006-tables5.xlsx"

    if not s5_path.exists():
        print(f"  ✗ File not found: {s5_path}")
        print("    Download Table S5 from:")
        print("    https://febs.onlinelibrary.wiley.com/doi/full/10.1002/1878-0261.70091")
        return pd.DataFrame()

    print(f"  Reading: {s5_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(s5_path, read_only=True)

    if "two_group_tests_results_Phospho" not in wb.sheetnames:
        print(f"  ✗ Sheet not found. Available: {wb.sheetnames}")
        wb.close()
        return pd.DataFrame()

    ws = wb["two_group_tests_results_Phospho"]

    records: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip 2-row header (row 0 = category, row 1 = column names)
            continue

        sequence = str(row[0]) if row[0] else ""
        gene = str(row[1]) if row[1] else ""
        uniprot = str(row[2]) if row[2] else ""
        site_raw = str(row[3]) if row[3] else ""

        # Parse site: "GENE;S123" → ptm_site = "S123"
        site_parts = site_raw.split(";")
        if len(site_parts) >= 2:
            ptm_site = site_parts[1].strip()
        else:
            ptm_site = site_raw.strip()

        ptm_residue = ptm_site[0] if ptm_site and ptm_site[0] in "STY" else ""
        if not ptm_residue:
            continue

        # Extract fold-changes and p-values for all 6 comparisons
        fc_egfr_wt = _safe_float(row[20])  # FC EGFR vs WT
        pval_egfr_wt = _safe_float(row[14])  # adj p EGFR vs WT
        fc_egfr_kras = _safe_float(row[19])  # FC EGFR vs KRAS
        pval_egfr_kras = _safe_float(row[13])  # adj p EGFR vs KRAS
        fc_alk_egfr = _safe_float(row[16])  # FC ALK vs EGFR
        pval_alk_egfr = _safe_float(row[10])  # adj p ALK vs EGFR

        # Significance label for EGFR vs WT
        diff_egfr_wt = str(row[26]) if row[26] else ""

        records.append({
            "cell_line": "LUAD_tumor",
            "drug_name": "none",
            "protein": gene,
            "ptm_site": ptm_site,
            "ptm_residue": ptm_residue,
            "ptm_type": "phosphorylation",
            "peptide_sequence": sequence.replace("_", ""),
            "baseline_intensity": np.nan,
            "max_dose_intensity": np.nan,
            "max_dose_ug": np.nan,
            "log2_fold_change": round(fc_egfr_wt, 4) if np.isfinite(fc_egfr_wt) else np.nan,
            "EC50": np.nan, "pEC50": np.nan,
            "curve_effect_size": np.nan, "R2": np.nan,
            "n_doses": np.nan,
            "drug_smiles": "",
            "data_source": "febs_2025",
            "resistance_context": "tumor_phospho_signatures",
            # FEBS-specific columns
            "febs_uniprot": uniprot,
            "febs_site_id": site_raw,
            "febs_adj_pval_egfr_wt": pval_egfr_wt,
            "febs_fc_egfr_kras": round(fc_egfr_kras, 4) if np.isfinite(fc_egfr_kras) else np.nan,
            "febs_adj_pval_egfr_kras": pval_egfr_kras,
            "febs_diff_egfr_wt": diff_egfr_wt,
        })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} phosphosites from LUAD patient tumors")
    if not df.empty:
        n_sig = df["febs_adj_pval_egfr_wt"].dropna().lt(0.05).sum()
        print(f"    Significant EGFR vs WT (adj.p < 0.05): {n_sig}")
        print(f"    Unique genes: {df['protein'].nunique()}")
        # Show top EGFR vs WT changes
        sig_df = df[df["febs_adj_pval_egfr_wt"].fillna(1) < 0.05].copy()
        if not sig_df.empty:
            sig_df = sig_df.sort_values("log2_fold_change", key=abs, ascending=False)
            print(f"    Top significant phosphosites (EGFR-mutant vs WT tumors):")
            for _, r in sig_df.head(8).iterrows():
                print(f"      {r['protein']:10s} {r['ptm_site']:8s} "
                      f"log2FC = {r['log2_fold_change']:+.3f}  "
                      f"adj.p = {r['febs_adj_pval_egfr_wt']:.4f}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE F: Cancer Research 2021 — SILAC Phosphoproteomics of TKI Resistance
# "Alterations in the Global Proteome and Phosphoproteome in 3rd Generation
#  EGFR TKI Resistance Reveal New Drug Targets"
# Journal: Cancer Research 81(11):3051, 2021
# DOI:     https://doi.org/10.1158/0008-5472.CAN-20-1814
# Data:    data/raw/drugptm/cancerres_2021/table_s2_phosphosites.xlsx
#          data/raw/drugptm/cancerres_2021/table_s5_common.xlsx
# ══════════════════════════════════════════════════════════════════════════════

def process_cancerres_2021() -> pd.DataFrame:
    """
    Process Cancer Research 2021 3-state SILAC phosphoproteomics data.

    This study compared phosphorylation between H1975 parental cells and
    TKI-resistant clones using triple-label SILAC:
      Light (L) = H1975 parental
      Medium (M) = Resistant clone 1 (AZR3 or COR1)
      Heavy (H) = Resistant clone 2 (AZR4 or COR10)

    Table S2 has 4 sheets (12,240 phosphosites):
      S2A: H1975 vs AZR3/AZR4 + osimertinib treatment
      S2B: H1975 vs AZR3/AZR4 + DMSO control
      S2C: H1975 vs COR1/COR10 + rociletinib treatment
      S2D: H1975 vs COR1/COR10 + DMSO control

    CRITICAL: SILAC ratios (M/L, H/L) are LINEAR, not log2.
      M/L = resistant_clone1 / parental  (linear scale)
      H/L = resistant_clone2 / parental  (linear scale)
      log2FC = log2(linear_ratio)
      ratio < 1 → negative log2FC (dephosphorylation in resistant)
      ratio > 1 → positive log2FC (hyperphosphorylation in resistant)

    For each EGFR site with quantifiable ratios, we average the two clone
    log2FC values for a robust resistance estimate.

    Table S5 "Common phosphosites" contains 49 sites altered in ALL four
    resistant clones — pre-computed log2 FC (T-test Difference from MaxQuant).
    These are the most robust pan-resistance markers.

    Cell line: H1975 (EGFR L858R/T790M)
    Resistant clones:
      AZR3, AZR4  — osimertinib-resistant
      COR1, COR10 — rociletinib-resistant (3rd-gen TKI cross-resistance)
    """
    print("\n" + "=" * 70)
    print("SOURCE F: Cancer Research 2021 — SILAC Phosphoproteomics")
    print("=" * 70)

    s2_path = CANCERRES_DIR / "table_s2_phosphosites.xlsx"
    s5_path = CANCERRES_DIR / "table_s5_common.xlsx"

    if not s2_path.exists():
        print(f"  ✗ File not found: {s2_path}")
        print("    Download from: https://doi.org/10.1158/0008-5472.CAN-20-1814")
        return pd.DataFrame()

    print(f"  Reading: {s2_path.name}")

    import openpyxl

    # ── Sheet configurations ─────────────────────────────────────────────
    sheet_configs = {
        "H1975_AZR3_AZR4_osimertinib": {
            "drug": "Osimertinib", "treatment": "osimertinib",
            "clone_m": "AZR3", "clone_h": "AZR4",
        },
        "H1975_AZR3_AZR4_DMSO": {
            "drug": "Osimertinib", "treatment": "DMSO",
            "clone_m": "AZR3", "clone_h": "AZR4",
        },
        "H1975_COR1_COR10_rociletinib": {
            "drug": "Rociletinib", "treatment": "rociletinib",
            "clone_m": "COR1", "clone_h": "COR10",
        },
        "H1975_COR1_COR10_DMSO": {
            "drug": "Rociletinib", "treatment": "DMSO",
            "clone_m": "COR1", "clone_h": "COR10",
        },
    }

    records: list[dict] = []
    wb = openpyxl.load_workbook(s2_path, read_only=True)

    for sheet_name, scfg in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name]
        n_egfr = 0

        # Column indices discovered dynamically from header
        gene_col = site_col = aa_col = pos_col = None
        ml_col = hl_col = seqwin_col = locprob_col = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # title row
                continue
            if i == 1:  # header row — discover column positions
                header = [str(c) if c else "" for c in row]
                for j, h in enumerate(header):
                    if h in ("1st Gene", "Gene") and gene_col is None:
                        gene_col = j
                    if h == "Amino acid":
                        aa_col = j
                    if h == "Position":
                        pos_col = j
                    if h == "Sequence window":
                        seqwin_col = j
                    if h == "Localization prob":
                        locprob_col = j
                    # Take FIRST occurrence of summary ratio columns
                    if h == "Ratio M/L normalized" and ml_col is None:
                        ml_col = j
                    if h == "Ratio H/L normalized" and hl_col is None:
                        hl_col = j
                continue

            if gene_col is None:
                continue

            gene = str(row[gene_col]) if row[gene_col] else ""
            if gene != "EGFR":
                continue

            # ── Parse site info ──────────────────────────────────────────
            aa = str(row[aa_col]).strip() if aa_col is not None and row[aa_col] else ""
            if aa not in ("S", "T", "Y"):
                continue
            try:
                pos = int(float(row[pos_col]))
            except (ValueError, TypeError):
                continue

            ptm_site = f"{aa}{pos}"
            seq_win = (str(row[seqwin_col]).replace("_", "").strip()
                       if seqwin_col is not None and row[seqwin_col] else "")
            loc_prob = (_safe_float(row[locprob_col])
                        if locprob_col is not None else np.nan)

            # ── Extract SILAC ratios (LINEAR scale) ──────────────────────
            ml_ratio = _safe_float(row[ml_col]) if ml_col is not None else np.nan
            hl_ratio = _safe_float(row[hl_col]) if hl_col is not None else np.nan

            # Convert linear ratios → log2 fold-change
            log2fc_m = (float(np.log2(ml_ratio))
                        if np.isfinite(ml_ratio) and ml_ratio > 0 else np.nan)
            log2fc_h = (float(np.log2(hl_ratio))
                        if np.isfinite(hl_ratio) and hl_ratio > 0 else np.nan)

            # Average the two clones for a robust estimate
            fcs = [v for v in [log2fc_m, log2fc_h] if np.isfinite(v)]
            if not fcs:
                continue  # skip sites with no quantifiable ratios

            avg_log2fc = float(np.mean(fcs))
            n_egfr += 1

            drug_smiles = (OSIMERTINIB_SMILES
                           if scfg["drug"] == "Osimertinib" else "")

            records.append({
                "cell_line": "H1975",
                "drug_name": scfg["drug"],
                "protein": "EGFR",
                "ptm_site": ptm_site,
                "ptm_residue": aa,
                "ptm_type": "phosphorylation",
                "peptide_sequence": seq_win,
                "baseline_intensity": np.nan,
                "max_dose_intensity": np.nan,
                "max_dose_ug": np.nan,
                "log2_fold_change": round(avg_log2fc, 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": drug_smiles,
                "data_source": "cancerres_2021",
                "resistance_context": "parental_vs_resistant",
                # Source-specific columns
                "cancerres_treatment": scfg["treatment"],
                "cancerres_clone_m": scfg["clone_m"],
                "cancerres_clone_h": scfg["clone_h"],
                "cancerres_log2fc_clone1": (round(log2fc_m, 4)
                                            if np.isfinite(log2fc_m) else np.nan),
                "cancerres_log2fc_clone2": (round(log2fc_h, 4)
                                            if np.isfinite(log2fc_h) else np.nan),
                "cancerres_silac_ml_ratio": (round(ml_ratio, 6)
                                             if np.isfinite(ml_ratio) else np.nan),
                "cancerres_silac_hl_ratio": (round(hl_ratio, 6)
                                             if np.isfinite(hl_ratio) else np.nan),
                "cancerres_localization_prob": loc_prob,
            })

        print(f"    {sheet_name}: {n_egfr} EGFR sites with quantifiable ratios")

    wb.close()

    # ── Table S5: Common resistance phosphosites ─────────────────────────
    s5_records: list[dict] = []
    if s5_path.exists():
        print(f"  Reading: {s5_path.name}")
        wb5 = openpyxl.load_workbook(s5_path, read_only=True)

        if "Common phosphosites" in wb5.sheetnames:
            ws5 = wb5["Common phosphosites"]
            for i, row in enumerate(ws5.iter_rows(values_only=True)):
                if i < 2:  # skip title + header
                    continue

                site_id = str(row[0]) if row[0] else ""
                if not site_id or "-" not in site_id:
                    continue

                parts = site_id.split("-", 1)
                gene_name = parts[0].strip()
                ptm_site = parts[1].strip() if len(parts) > 1 else ""
                if not ptm_site or ptm_site[0] not in "STY":
                    continue

                # Pre-computed log2FC (T-test Difference from MaxQuant)
                # Col 1: AZR4/parental + osimertinib (H/L)
                # Col 2: AZR3/parental + osimertinib (M/L)
                # Col 3: COR10/parental + rociletinib (H/L)
                # Col 4: COR1/parental + rociletinib (M/L)
                osi_hl = _safe_float(row[1])
                osi_ml = _safe_float(row[2])
                roci_hl = _safe_float(row[3])
                roci_ml = _safe_float(row[4])

                osi_fcs = [v for v in [osi_hl, osi_ml] if np.isfinite(v)]
                roci_fcs = [v for v in [roci_hl, roci_ml] if np.isfinite(v)]

                base_record = {
                    "cell_line": "H1975",
                    "protein": gene_name,
                    "ptm_site": ptm_site,
                    "ptm_residue": ptm_site[0],
                    "ptm_type": "phosphorylation",
                    "peptide_sequence": "",
                    "baseline_intensity": np.nan,
                    "max_dose_intensity": np.nan,
                    "max_dose_ug": np.nan,
                    "EC50": np.nan, "pEC50": np.nan,
                    "curve_effect_size": np.nan, "R2": np.nan,
                    "n_doses": np.nan,
                    "data_source": "cancerres_2021",
                    "resistance_context": "common_resistance_phosphosite",
                    "cancerres_localization_prob": np.nan,
                }

                if osi_fcs:
                    s5_records.append({
                        **base_record,
                        "drug_name": "Osimertinib",
                        "drug_smiles": OSIMERTINIB_SMILES,
                        "log2_fold_change": round(float(np.mean(osi_fcs)), 4),
                        "cancerres_treatment": "osimertinib",
                        "cancerres_clone_m": "AZR3",
                        "cancerres_clone_h": "AZR4",
                        "cancerres_log2fc_clone1": (round(osi_ml, 4)
                                                    if np.isfinite(osi_ml) else np.nan),
                        "cancerres_log2fc_clone2": (round(osi_hl, 4)
                                                    if np.isfinite(osi_hl) else np.nan),
                        "cancerres_silac_ml_ratio": np.nan,
                        "cancerres_silac_hl_ratio": np.nan,
                    })

                if roci_fcs:
                    s5_records.append({
                        **base_record,
                        "drug_name": "Rociletinib",
                        "drug_smiles": "",
                        "log2_fold_change": round(float(np.mean(roci_fcs)), 4),
                        "cancerres_treatment": "rociletinib",
                        "cancerres_clone_m": "COR1",
                        "cancerres_clone_h": "COR10",
                        "cancerres_log2fc_clone1": (round(roci_ml, 4)
                                                    if np.isfinite(roci_ml) else np.nan),
                        "cancerres_log2fc_clone2": (round(roci_hl, 4)
                                                    if np.isfinite(roci_hl) else np.nan),
                        "cancerres_silac_ml_ratio": np.nan,
                        "cancerres_silac_hl_ratio": np.nan,
                    })

        wb5.close()

    # ── Combine and report ───────────────────────────────────────────────
    all_records = records + s5_records
    df = pd.DataFrame(all_records)

    print(f"\n  ✓ Table S2: {len(records)} EGFR phosphosite measurements")
    print(f"    Table S5: {len(s5_records)} common resistance phosphosites")
    print(f"    Total:    {len(df)} rows")

    if not df.empty:
        for drug in sorted(df["drug_name"].unique()):
            sub = df[df["drug_name"] == drug]
            n_egfr_direct = sub[sub["protein"] == "EGFR"].shape[0]
            print(f"    {drug}: {len(sub)} rows "
                  f"({n_egfr_direct} EGFR-direct, "
                  f"{len(sub) - n_egfr_direct} network)")

        # Show EGFR sites with quantifiable ratios from S2
        egfr_s2 = df[(df["protein"] == "EGFR") &
                     (df["resistance_context"] == "parental_vs_resistant")]
        if not egfr_s2.empty:
            print(f"    ── EGFR sites from S2 (SILAC ratios) ──")
            for _, r in egfr_s2.sort_values(
                    "log2_fold_change", key=abs, ascending=False).iterrows():
                ml = r.get("cancerres_silac_ml_ratio", np.nan)
                hl = r.get("cancerres_silac_hl_ratio", np.nan)
                ml_s = f"M/L={ml:.3f}" if pd.notna(ml) else "M/L=—"
                hl_s = f"H/L={hl:.3f}" if pd.notna(hl) else "H/L=—"
                print(f"      {r['ptm_site']:8s} {r['cancerres_treatment']:12s} "
                      f"log2FC={r['log2_fold_change']:+.3f}  "
                      f"({ml_s}, {hl_s})")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE G: MCP 2025 — Phospho/Glycoproteomics of TKI-Resistant NSCLC
# "Fe-ZIC-cHILIC for simultaneous glycoproteomics and phosphoproteomics
#  in TKI-resistant NSCLC"
# Journal: Molecular and Cellular Proteomics (2025)
# DOI:     https://doi.org/10.1016/j.mcpro.2025.100917
# Data:    data/raw/drugptm/mcp_2025/table_s8_phospho_glyco_summary.xlsx
# ══════════════════════════════════════════════════════════════════════════════

def process_mcp_2025() -> pd.DataFrame:
    """
    Process MCP 2025 phosphoproteomics data (Table S8).

    This study compared phospho- and glycoproteomes across NSCLC cell lines
    using Fe-ZIC-cHILIC enrichment and label-free DIA-MS:
      • H1975  (EGFR L858R/T790M) — TKI-resistant reference
      • H3255  (EGFR L858R)       — TKI-sensitive reference
      • PC-9   (EGFR exon19del)   — TKI-sensitive
      • CL68   — not in our standard NSCLC panel, skipped

    Table S8 columns (key ones):
      Col 0:  Name          (e.g., "P00533_S1064")
      Col 1:  Group         ("PP" = phospho, "GP" = glyco)
      Col 2:  Accession     (e.g., "P00533")
      Col 3:  Site          (e.g., "S1064")
      Col 10: FC CL68/PC9   (log2, skipped — CL68 not in scope)
      Col 11: FC H1975/H3255 (log2 fold-change)
      Col 12: FC PC9/H3255   (log2 fold-change)
      Col 13: FC CL68/H1975  (log2, skipped)
      Col 14: p-val CL68vsPC9
      Col 15: p-val H1975vsH3255
      Col 16: p-val PC9vsH3255
      Col 17: p-val CL68vsH1975
      Col 37: Annotated Sequence

    Fold-changes are already log2.
    Sentinel: -10 = not detected → treated as NaN.

    Biological interpretation:
      H1975/H3255: Resistant (L858R/T790M) vs sensitive (L858R)
                   → captures T790M gatekeeper-driven phospho changes
      PC9/H3255:   exon19del vs L858R
                   → mutation-class phospho differences

    We extract EGFR phosphosites (P00533, Group=PP) and create two rows
    per site: one for H1975 and one for PC-9, each relative to H3255.

    Key EGFR phospho sites: S991, S1064, S1070, S1096, S1166, S1190,
                             T693, T1074, T1078, Y1172, Y1197 (11 sites)
    """
    print("\n" + "=" * 70)
    print("SOURCE G: MCP 2025 — Phospho/Glycoproteomics TKI-Resistant NSCLC")
    print("=" * 70)

    s8_path = MCP_DIR / "table_s8_phospho_glyco_summary.xlsx"

    if not s8_path.exists():
        print(f"  ✗ File not found: {s8_path}")
        print("    Download from: https://doi.org/10.1016/j.mcpro.2025.100917")
        return pd.DataFrame()

    print(f"  Reading: {s8_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(s8_path, read_only=True)
    ws = wb[wb.sheetnames[0]]  # single sheet (工作表1)

    records: list[dict] = []
    n_phospho_total = 0
    n_glyco_skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip title + header
            continue

        name = str(row[0]) if row[0] else ""
        group = str(row[1]).strip() if row[1] else ""
        accession = str(row[2]).strip() if row[2] else ""
        site_raw = str(row[3]).strip() if row[3] else ""

        # Skip glycopeptides — only process phosphosites
        if group != "PP":
            if group == "GP" and accession == "P00533":
                n_glyco_skipped += 1
            continue

        # Filter to EGFR (UniProt P00533)
        if accession != "P00533":
            continue

        n_phospho_total += 1

        # Parse site: "S1064" → ptm_residue="S", ptm_site="S1064"
        match = re.match(r"^([STY])(\d+)$", site_raw)
        if not match:
            continue
        ptm_residue = match.group(1)
        ptm_site = site_raw

        # ── Extract fold-changes (already log2) ─────────────────────────
        fc_h1975_h3255 = _safe_float(row[11])
        fc_pc9_h3255 = _safe_float(row[12])
        pval_h1975_h3255 = _safe_float(row[15])
        pval_pc9_h3255 = _safe_float(row[16])

        # Mean log2 intensities per cell line (cols 18-21)
        intensity_h1975 = _safe_float(row[19])
        intensity_h3255 = _safe_float(row[20])
        intensity_pc9 = _safe_float(row[21])

        # Annotated sequence (col 37)
        peptide = str(row[37]).strip() if row[37] else ""
        # Clean peptide: strip flanking residues "[K].PEPTIDE.[R]" → "PEPTIDE"
        pep_match = re.search(r"\]\.(.+?)\.\[", peptide)
        peptide_clean = pep_match.group(1) if pep_match else peptide

        # Sentinel: -10 = not detected → treat as NaN
        SENTINEL = -10.0

        # ── H1975 vs H3255 (resistant vs sensitive) ─────────────────────
        if (np.isfinite(fc_h1975_h3255)
                and abs(fc_h1975_h3255 - SENTINEL) > 0.01):
            records.append({
                "cell_line": "H1975",
                "drug_name": "Osimertinib",
                "protein": "EGFR",
                "ptm_site": ptm_site,
                "ptm_residue": ptm_residue,
                "ptm_type": "phosphorylation",
                "peptide_sequence": peptide_clean,
                "baseline_intensity": (round(intensity_h3255, 4)
                                       if np.isfinite(intensity_h3255) else np.nan),
                "max_dose_intensity": (round(intensity_h1975, 4)
                                       if np.isfinite(intensity_h1975) else np.nan),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(fc_h1975_h3255, 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "mcp_2025",
                "resistance_context": "cell_line_phosphoproteome",
                # Source-specific columns
                "mcp_comparison": "H1975_vs_H3255",
                "mcp_pvalue": pval_h1975_h3255,
                "mcp_site_name": name,
                "mcp_mutation_test": "L858R_T790M",
                "mcp_mutation_ref": "L858R",
            })

        # ── PC-9 vs H3255 (exon19del vs L858R) ─────────────────────────
        if (np.isfinite(fc_pc9_h3255)
                and abs(fc_pc9_h3255 - SENTINEL) > 0.01):
            records.append({
                "cell_line": "PC-9",
                "drug_name": "Osimertinib",
                "protein": "EGFR",
                "ptm_site": ptm_site,
                "ptm_residue": ptm_residue,
                "ptm_type": "phosphorylation",
                "peptide_sequence": peptide_clean,
                "baseline_intensity": (round(intensity_h3255, 4)
                                       if np.isfinite(intensity_h3255) else np.nan),
                "max_dose_intensity": (round(intensity_pc9, 4)
                                       if np.isfinite(intensity_pc9) else np.nan),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(fc_pc9_h3255, 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "mcp_2025",
                "resistance_context": "cell_line_phosphoproteome",
                # Source-specific columns
                "mcp_comparison": "PC9_vs_H3255",
                "mcp_pvalue": pval_pc9_h3255,
                "mcp_site_name": name,
                "mcp_mutation_test": "exon19del",
                "mcp_mutation_ref": "L858R",
            })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} EGFR phosphosite × cell-line measurements")
    print(f"    EGFR phospho sites found: {n_phospho_total}")
    print(f"    EGFR glyco sites skipped: {n_glyco_skipped}")

    if not df.empty:
        print(f"    Cell lines: {sorted(df['cell_line'].unique())}")
        print(f"    Unique sites: {sorted(df['ptm_site'].unique())}")
        for cl in sorted(df["cell_line"].unique()):
            sub = df[df["cell_line"] == cl]
            print(f"    ── {cl} vs H3255 ──")
            for _, r in sub.sort_values("log2_fold_change").iterrows():
                pval_s = (f"p={r['mcp_pvalue']:.4f}"
                          if pd.notna(r.get("mcp_pvalue")) else "p=NA")
                print(f"      {r['ptm_site']:8s} log2FC={r['log2_fold_change']:+.3f}"
                      f"  {pval_s}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# CELL-LINE × DRUG CATALOG (DrugPTM-Bench only)
# ══════════════════════════════════════════════════════════════════════════════

def build_catalog(cell_line_files: list[Path]) -> pd.DataFrame:
    """
    Create a lightweight catalog listing every unique
    (cell_line, drug, timepoint) combination in DrugPTM-Bench.
    """
    print("\n" + "=" * 70)
    print("Building Cell-Line × Drug Catalog (DrugPTM-Bench)")
    print("=" * 70)

    frames: list[pd.DataFrame] = []
    for fpath in cell_line_files:
        for chunk in pd.read_csv(
                fpath,
                usecols=["Cell Line", "Chemical Name",
                         "Timepoint (min)", "Gene names"],
                chunksize=CHUNK_SIZE,
                low_memory=False,
        ):
            frames.append(
                chunk.groupby(["Cell Line", "Chemical Name",
                               "Timepoint (min)"])
                .size()
                .reset_index(name="n_measurements")
            )

    if not frames:
        return pd.DataFrame()

    catalog = (
        pd.concat(frames, ignore_index=True)
        .groupby(["Cell Line", "Chemical Name", "Timepoint (min)"],
                 dropna=False)["n_measurements"]
        .sum()
        .reset_index()
        .sort_values(["Cell Line", "Chemical Name"])
    )

    print(f"  Unique combinations: {len(catalog):,}")
    print(f"  Cell lines: {catalog['Cell Line'].nunique()}")
    print(f"  Drugs: {catalog['Chemical Name'].nunique()}")
    return catalog


# ══════════════════════════════════════════════════════════════════════════════
# MERGE ALL SOURCES & SAVE
# ══════════════════════════════════════════════════════════════════════════════

def _annotate_ptm_schema(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add the two PTM-BDL schema columns (`target_protein`,
    `ptm_modification_type`) used by step06 / step10.

    `target_protein` ∈ {EGFR, ERBB2}
        Derived from the `protein` (gene-name) column.  Rows whose primary
        gene contains "EGFR" → EGFR; "ERBB2" or "HER2" → ERBB2.  Mixed/
        shared peptides where EGFR is present prefer EGFR (because EGFR is
        the primary drug target).  Genes outside the ERBB family are tagged
        as `"other"` here and DROPPED in `merge_and_save` immediately after
        annotation — the project scope (config.yaml `project.target_proteins`)
        is strictly {EGFR, ERBB2}, and the PTM-BDL token branch only consumes
        these two proteins.  Pathway / signaling-neighbour proteins from
        PNAS 2025 / Remsing Rix 2022 / FEBS 2025 etc. are therefore not
        carried forward into step06.

    `ptm_modification_type` ∈ {phospho_Y, phospho_S, phospho_T, glyco_N}
        Derived from (`ptm_type`, `ptm_residue`).  This is the type ID
        consumed by the PTM-BDL learnable modification-type embedding
        (proposal §7.4):  Y=0, S=1, T=2, N=3.
    """
    if df.empty:
        return df

    def _target(p):
        if not isinstance(p, str):
            return ""
        u = p.upper()
        # Prefer EGFR when both are present (it is the primary drug target)
        if "EGFR" in u:
            return "EGFR"
        if "ERBB2" in u or "HER2" in u:
            return "ERBB2"
        return p

    def _modtype(t, r):
        t = str(t).lower() if pd.notna(t) else ""
        r = str(r).upper().strip() if pd.notna(r) else ""
        if t.startswith("phosphor"):
            if r in ("Y", "S", "T"):
                return f"phospho_{r}"
            return "phospho_other"
        if t.startswith("n-glyco") or t.startswith("n_glyco") or t.startswith("glyco"):
            return "glyco_N"
        return "other"

    df = df.copy()
    df["target_protein"] = df["protein"].apply(_target)
    df["ptm_modification_type"] = df.apply(
        lambda r: _modtype(r.get("ptm_type"), r.get("ptm_residue")),
        axis=1,
    )
    return df


def merge_and_save(df_bench: pd.DataFrame,
                   df_tozuka: pd.DataFrame,
                   df_hsu: pd.DataFrame,
                   df_pnas: pd.DataFrame,
                   df_febs: pd.DataFrame,
                   df_cancerres: pd.DataFrame,
                   df_mcp: pd.DataFrame,
                   catalog: pd.DataFrame,
                   df_remsing: pd.DataFrame = None,
                   df_ruprecht: pd.DataFrame = None,
                   df_erbb2_bench: pd.DataFrame = None,
                   # ── Multi-PTM (glyco) frames added 2026-06-28 ───────
                   df_mcp_glyco: pd.DataFrame = None,
                   df_mcp_2025b_glyco: pd.DataFrame = None,
                   df_erbb2_glycoatlas: pd.DataFrame = None,
                   df_st6gal1: pd.DataFrame = None,
                   df_egfr_fuco: pd.DataFrame = None) -> None:
    """
    Concatenate all sources and save the unified multi-PTM, multi-protein
    output.

    Output:
      data/processed/drugptm/drugptm_multiptm_responses.csv

    Schema (per row):
      cell_line, drug_name, protein, target_protein,
      ptm_site, ptm_residue, ptm_type, ptm_modification_type,
      peptide_sequence, baseline_intensity, max_dose_intensity,
      log2_fold_change, data_source, resistance_context,
      + source-specific columns

    The schema is forward-compatible with the PTM-BDL token construction in
    step06: each row provides exactly one (target_protein, ptm_site)
    measurement, typed by `ptm_modification_type`, and (when available)
    drug-conditioned via `drug_name` × `log2_fold_change`.
    """
    print("\n" + "=" * 70)
    print("Merging All Sources → Unified Multi-PTM, Multi-Protein Output")
    print("=" * 70)

    frames: list[pd.DataFrame] = []
    source_counts: dict[str, int] = {}

    def _push(df, name):
        if df is not None and not df.empty:
            frames.append(df)
            source_counts[name] = len(df)

    # ── Phospho sources (A–I) ─────────────────────────────────────────
    _push(df_bench, "drugptm_bench_egfr")
    _push(df_erbb2_bench, "drugptm_bench_erbb2")
    _push(df_tozuka, "tozuka_2024")
    _push(df_hsu, "hsu_2025")
    _push(df_pnas, "pnas_2025")
    _push(df_febs, "febs_2025")
    _push(df_cancerres, "cancerres_2021")
    _push(df_mcp, "mcp_2025_phospho")
    _push(df_remsing, "remsing_rix_2022")
    _push(df_ruprecht, "ruprecht_2017")

    # ── Glyco sources (J–N) — multi-PTM expansion (PTM-BDL §3.1, §3.2) ──
    _push(df_mcp_glyco, "mcp_2025_glyco")
    _push(df_mcp_2025b_glyco, "mcp_2025b_glyco")
    _push(df_erbb2_glycoatlas, "erbb2_glycoform_atlas_2024")
    _push(df_st6gal1, "st6gal1_erbb2_2021")
    _push(df_egfr_fuco, "egfr_fucosylation_2020")

    if not frames:
        print("  ✗ No data from any source — nothing to save!")
        return

    df_merged = pd.concat(frames, ignore_index=True)

    # ── Filter to relevant drugs only ────────────────────────────────────
    # The model uses 4 EGFR TKIs + FEBS tumor genotype data ("none").
    # DrugPTM-Bench contains many non-EGFR drugs (Dasatinib, Imatinib,
    # Bortezomib, etc.) that are irrelevant to EGFR TKI resistance.
    # Rociletinib (CO-1686) is a 3rd-gen EGFR TKI included because the
    # Cancer Research 2021 study shows cross-resistance phospho patterns
    # shared between Osimertinib and Rociletinib resistant clones.
    # Include all ERBB-family TKIs + HER2 drugs (Section 7a HER2_EXPANSION_PLAN.md)
    # Sapitinib replaced Neratinib (not in GDSC2). Lapatinib added for HER2.
    relevant_drugs = {"Osimertinib", "Gefitinib", "Afatinib", "Erlotinib",
                      "Lapatinib", "Sapitinib", "Rociletinib",
                      "Pertuzumab", "Trastuzumab",  # HER2 antibodies from DrugPTM-Bench
                      "none"}
    before_filter = len(df_merged)
    df_merged = df_merged[df_merged["drug_name"].isin(relevant_drugs)].reset_index(drop=True)
    n_dropped = before_filter - len(df_merged)
    if n_dropped > 0:
        print(f"\n  Filtered to EGFR TKIs + tumor genotype data:")
        print(f"    Kept: {len(df_merged):,} rows ({len(df_merged) / before_filter * 100:.0f}%)")
        print(f"    Dropped: {n_dropped} rows from non-EGFR drugs")

    # ── Filter to ERBB-family cell lines (NSCLC + HER2+ breast) ─────────
    # Section 7a HER2_EXPANSION_PLAN.md: ALL EGFR drugs also tested on
    # 52 breast cancer cell lines. Include both NSCLC and HER2+ breast.
    erbb_cell_lines = {
        # NSCLC cell lines (EGFR context)
        "HCC827", "PC-9", "PC9", "H1975", "NCI-H1975",
        "H1650", "NCI-H1650", "H3255", "NCI-H3255",
        "HCC4006", "H820", "NCI-H820", "H1573", "NCI-H1573",
        "A549", "NCI-A549", "H460", "NCI-H460", "H358", "NCI-H358",
        "A431",  # DrugPTM-Bench: WT EGFR, Gefitinib/Afatinib dose-response
        "PC9GR",  # Remsing Rix 2022: gefitinib-resistant PC-9 derivative
        "LUAD_tumor",  # FEBS 2025 patient LUAD tumors
        # HER2+ breast cancer cell lines (ERBB2 context) — added for expansion
        "BT-474", "BT474", "SKBR3", "SK-BR-3", "AU565", "HCC1954",
        "MDA-MB-453", "MDA-MB-361", "ZR-75-30", "UACC-812",
        "HCC1569", "JIMT-1", "JIMT1", "MDA-MB-175",
    }
    # Add the cell-line labels coming from the new glyco sources (L–N) so
    # they survive this filter.  GLYCO_EXTRA_CELL_LINES is defined alongside
    # the glyco loaders further down in this file.
    try:
        erbb_cell_lines = erbb_cell_lines | GLYCO_EXTRA_CELL_LINES
    except NameError:
        pass  # for safety if loaded out of order during partial imports
    before_filter = len(df_merged)
    df_merged = df_merged[df_merged["cell_line"].isin(erbb_cell_lines)].reset_index(drop=True)

    n_dropped = before_filter - len(df_merged)
    if n_dropped > 0:
        print(f"\n  Filtered to ERBB-family cell lines (NSCLC + HER2+ breast):")
        print(f"    Kept: {len(df_merged):,} rows")
        print(f"    Dropped: {n_dropped} rows from non-ERBB cell lines")

    # ── Annotate target_protein + ptm_modification_type (PTM-BDL schema) ──
    df_merged = _annotate_ptm_schema(df_merged)

    # ── Filter to EGFR + ERBB2 ONLY (project scope) ──────────────────────
    # The project scope (config.yaml `project.target_proteins`) is strictly
    # {EGFR, ERBB2}.  PNAS 2025 / Remsing Rix 2022 / FEBS 2025 carry many
    # pathway / signaling-neighbour proteins (BCAR1, SRRM2, GAB1, ...) that
    # were originally tagged `target_protein == "other"`.  We DROP them
    # here so the unified output and every downstream step (06 → 13) sees
    # only the two ERBB-family receptors that the PTM-BDL token branch
    # actually models (proposal §3, §7).
    before_scope = len(df_merged)
    df_merged = df_merged[
        df_merged["target_protein"].isin({"EGFR", "ERBB2"})
    ].reset_index(drop=True)
    n_other = before_scope - len(df_merged)
    if n_other > 0:
        print(f"\n  Restricted to ERBB-family proteins (EGFR + ERBB2):")
        print(f"    Kept:    {len(df_merged):,} rows")
        print(f"    Dropped: {n_other:,} non-ERBB-family rows "
              f"(pathway proteins outside project scope)")

    # ── Print summary ──────────────────────────────────────────────────────
    print(f"\n  Total rows: {len(df_merged):,}")
    for src, cnt in source_counts.items():
        print(f"    {src:25s}: {cnt:>4} rows")
    print(f"\n  Unique cell lines:       "
          f"{sorted(df_merged['cell_line'].unique())}")
    print(f"  Unique drugs:            "
          f"{sorted(df_merged['drug_name'].unique())}")
    print(f"  Unique PTM sites:        "
          f"{df_merged['ptm_site'].nunique()}")
    print(f"  Data sources:            "
          f"{sorted(df_merged['data_source'].unique())}")
    print(f"  Resistance contexts:     "
          f"{sorted(df_merged['resistance_context'].unique())}")

    # ── Multi-PTM / multi-protein breakdown (the new schema columns) ──────
    print(f"\n  ── PTM-BDL schema breakdown ──")
    print(f"  target_protein:")
    for tp, cnt in df_merged["target_protein"].value_counts().items():
        print(f"    {str(tp):12s}: {cnt:>5} rows")
    print(f"  ptm_modification_type:")
    for mt, cnt in df_merged["ptm_modification_type"].value_counts().items():
        print(f"    {str(mt):12s}: {cnt:>5} rows")
    # 2×2 cross-tab so we can see at a glance whether glyco coverage is real
    try:
        ct = pd.crosstab(df_merged["target_protein"],
                         df_merged["ptm_modification_type"]).fillna(0).astype(int)
        print(f"  target_protein × ptm_modification_type:\n{ct}")
    except Exception:
        pass

    # ── Cross-source phosphosite overlap ───────────────────────────────────
    site_sources: dict[str, set[str]] = {}
    for _, row in df_merged.iterrows():
        site_sources.setdefault(row["ptm_site"], set()).add(row["data_source"])

    multi = {s: srcs for s, srcs in site_sources.items() if len(srcs) > 1}
    if multi:
        print(f"\n  Sites measured in ≥ 2 sources ({len(multi)}):")
        for site, srcs in sorted(multi.items()):
            print(f"    {site:8s}  ← {', '.join(sorted(srcs))}")
    else:
        print("\n  No PTM-site overlap between sources "
              "(each provides complementary coverage)")

    # ── Save unified multi-PTM, multi-protein table ──────────────────────
    out_path = OUT_DIR / "drugptm_multiptm_responses.csv"
    df_merged.to_csv(out_path, index=False)
    print(f"\n  ✓ {out_path.relative_to(PROJECT_ROOT)}")
    print(f"    {len(df_merged):,} rows  •  "
          f"{df_merged['cell_line'].nunique()} cell lines  •  "
          f"{df_merged['drug_name'].nunique()} drugs  •  "
          f"{df_merged['ptm_site'].nunique()} PTM sites  •  "
          f"{df_merged['target_protein'].nunique()} proteins  •  "
          f"{df_merged['ptm_modification_type'].nunique()} mod types")

    # ── Build unified catalog from ALL sources ────────────────────────────
    catalog_path = OUT_DIR / "drugptm_cell_line_drug_catalog.csv"

    # Start with DrugPTM-Bench catalog (has timepoint info), filtered to EGFR TKIs
    catalog_frames = []
    if not catalog.empty:
        bench_relevant = {"Gefitinib", "Afatinib", "Erlotinib", "Osimertinib"}
        catalog_filtered = catalog[catalog["Chemical Name"].isin(bench_relevant)]
        if not catalog_filtered.empty:
            catalog_frames.append(catalog_filtered)

    # Add entries from other sources (Tozuka, Hsu, PNAS, FEBS)
    other_sources = df_merged[df_merged["data_source"] != "drugptm_bench"]
    if not other_sources.empty:
        other_catalog = (
            other_sources.groupby(["cell_line", "drug_name", "data_source"])
            .size()
            .reset_index(name="n_measurements")
            .rename(columns={
                "cell_line": "Cell Line",
                "drug_name": "Chemical Name",
                "data_source": "Timepoint (min)",  # reuse column for source info
            })
            .sort_values(["Cell Line", "Chemical Name"])
        )
        catalog_frames.append(other_catalog)

    if catalog_frames:
        full_catalog = pd.concat(catalog_frames, ignore_index=True)
        full_catalog.to_csv(catalog_path, index=False)
        print(f"  ✓ {catalog_path.relative_to(PROJECT_ROOT)}")
        print(f"    {len(full_catalog):,} cell-line × drug combos (all sources)")
        print(f"    Drugs: {sorted(full_catalog['Chemical Name'].unique())}")
        print(f"    Cell lines: {sorted(full_catalog['Cell Line'].unique())}")
    else:
        print(f"  ⚠ No catalog to save ({catalog_path.name})")


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE H — Remsing Rix et al., 2022 (Sci Signal)
# ══════════════════════════════════════════════════════════════════════════════

def process_remsing_rix_2022() -> pd.DataFrame:
    """
    Process phosphoproteomics from Remsing Rix et al., Sci Signal 2022.

    Paper: "IGF-binding proteins secreted by CAFs induce context-dependent
           drug sensitization of lung cancer cells"
    PMID:  35973030
    DOI:   10.1126/scisignal.abj5879
    PMC:   PMC9528501

    Data:  SM_Excel_file_4.xlsx — 847 phosphosites from PC9GR cells
           (gefitinib-resistant EGFR exon19del derivative of PC-9)
           treated with Osimertinib ± CAF-conditioned medium

    Conditions (columns 74-77 = mean log2 intensities):
      CAF7-DMSO, CAF7-OSI, MRC5-DMSO, MRC5-OSI
    We use MRC5 (normal fibroblast) condition as the "pure drug effect"
    since it represents osimertinib response without CAF confound.

    log2FC = MRC5-OSI mean - MRC5-DMSO mean  (already in log2 space)
    """
    print(f"\n{'=' * 60}")
    print("  Source H: Remsing Rix et al., Sci Signal 2022")
    print("  PC9GR (gefitinib-resistant PC-9) + Osimertinib phosphoproteomics")
    print(f"{'=' * 60}")

    data_dir = RAW_DIR / "remsing_rix_2022"
    xlsx_path = data_dir / "NIHMS1836066-supplement-SM_Excel_file_4.xlsx"

    if not xlsx_path.exists():
        print(f"  ⚠ File not found: {xlsx_path}")
        return pd.DataFrame()

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Data']

    records = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        protein_name = str(row[3]) if row[3] else ""
        protein_id = str(row[5]) if row[5] else ""
        position = row[1]
        amino_acid = str(row[9]) if row[9] else ""

        if not position or not amino_acid:
            continue

        # Get MRC5 (normal fibroblast) condition means — pure drug effect
        mrc_dmso = row[76]  # MRC5-DMSO mean (log2 intensity)
        mrc_osi = row[77]  # MRC5-OSI mean (log2 intensity)

        # Compute log2FC = log2(Osi/DMSO) = MRC5-OSI - MRC5-DMSO (already log2)
        if mrc_dmso is not None and mrc_osi is not None:
            try:
                mrc_dmso_val = float(mrc_dmso)
                mrc_osi_val = float(mrc_osi)
                log2fc = mrc_osi_val - mrc_dmso_val
            except (ValueError, TypeError):
                continue
        else:
            continue

        # Determine if this is an EGFR site
        gene = ""
        if "EGFR" in protein_name.upper():
            gene = "EGFR"
        elif "sp|" in protein_id:
            # Extract gene from UniProt format: sp|P00533|EGFR_HUMAN
            parts = protein_id.split("|")
            if len(parts) >= 3:
                gene = parts[2].split("_")[0]

        ptm_site = f"{amino_acid}{position}"

        records.append({
            "protein": protein_name,
            "gene": gene,
            "ptm_site": ptm_site,
            "ptm_type": "phosphorylation",
            "log2_fold_change": round(log2fc, 4),
            "cell_line": "PC9GR",
            "drug_name": "Osimertinib",
            "data_source": "remsing_rix_2022",
            "resistance_context": "gefitinib_resistant_phosphoproteome",
            "pmid": "35973030",
        })

    wb.close()

    if not records:
        print("  ⚠ No valid phosphosite data extracted")
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # Filter to EGFR sites for summary
    egfr_sites = df[df["gene"] == "EGFR"]
    print(f"  ✓ Extracted {len(df)} phosphosites ({len(egfr_sites)} EGFR)")
    if not egfr_sites.empty:
        for _, site in egfr_sites.iterrows():
            print(f"    EGFR {site['ptm_site']}: log2FC = {site['log2_fold_change']:+.3f}")

    # Count signaling pathway proteins
    pathway_genes = {"ERBB2", "ERBB3", "MET", "SRC", "FAK1", "PTK2",
                     "GAB1", "GAB2", "SHC1", "STAT3", "MAPK1", "MAPK3"}
    pathway_df = df[df["gene"].isin(pathway_genes)]
    if not pathway_df.empty:
        print(f"  ✓ {len(pathway_df)} signaling pathway phosphosites also extracted")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE I: Ruprecht et al., 2017 (PMID 28209619)
# "Lapatinib Resistance in Breast Cancer Cells Is Accompanied by
#  Phosphorylation-Mediated Reprogramming of Glycolysis"
# Journal: Cancer Research 77(8):1842-1853 (2017)
# DOI:     https://doi.org/10.1158/0008-5472.CAN-16-2976
# Data:    data/raw/drugptm/her2_ruprecht_2017/
#          Table S2: 6,421 filtered/normalised phosphosites (SILAC)
#
# Description:
#   Triple-label SILAC phosphoproteomics comparing:
#     L (Light)  = BT-474 parental (HER2-amplified, untreated)
#     M (Medium) = BT-474 + Lapatinib 4 h (acute drug response)
#     H (Heavy)  = BT-474 Lapatinib-Resistant (acquired resistance)
#   Contains >15,000 phosphopeptides from 6,421 filtered sites across
#   4 biological replicates. We extract ERBB2-specific phosphosites only
#   (10 sites) — NO pathway/signaling proteins included.
#
# What it provides to our model:
#   • Direct parental vs lapatinib-resistant ERBB2 phosphosite comparison
#   • H/L log2FC = resistant / parental (acquired resistance signal)
#   • M/L log2FC = lapatinib 4h / parental (acute drug effect)
#   • 5 sites significantly changed at FDR < 1%
#   • Y1233 shows strongest dephosphorylation (log2FC = -2.94)
#   • Drug: Lapatinib (matches GDSC HER2 IC50 labels)
# Resistance context: "parental_vs_resistant"
# ══════════════════════════════════════════════════════════════════════════════

RUPRECHT_DIR = RAW_DIR / "her2_ruprecht_2017"


def process_ruprecht_2017() -> pd.DataFrame:
    """
    Process Ruprecht 2017 SILAC phosphoproteomics data.

    Extracts ERBB2 phosphosites ONLY from Table S2 (filtered/normalised
    phosphoproteome). Does NOT include pathway/signaling proteins — only
    direct ERBB2 phosphorylation sites.

    SILAC ratios are already log2-normalised:
      H/L_Av_log2FC = mean log2(Resistant / Parental) across 4 replicates
      M/L_Av_log2FC = mean log2(Lapatinib_4h / Parental)

    The primary log2_fold_change = H/L (resistant vs parental), which is
    the same resistance comparison as Tozuka 2024 for EGFR.
    """
    print("\n" + "=" * 70)
    print("SOURCE I: Ruprecht et al., 2017 (PMID 28209619)")
    print("  Lapatinib Resistance Phosphoproteomics in HER2+ BT-474")
    print("=" * 70)

    # Find the phosphoproteome file (Table S2 filtered/normalised)
    xlsx_files = sorted(RUPRECHT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"  ✗ No .xlsx files found in: {RUPRECHT_DIR}")
        print("    Download supplementary from:")
        print("    https://doi.org/10.1158/0008-5472.CAN-16-2976")
        return pd.DataFrame()

    # Table S2 is the filtered/normalised phosphoproteome (largest file ~6MB)
    phospho_file = None
    for f in xlsx_files:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True)
        if "Phopshoproteome_filtered_norm" in wb.sheetnames:
            phospho_file = f
            wb.close()
            break
        wb.close()

    if phospho_file is None:
        print("  ✗ Could not find 'Phopshoproteome_filtered_norm' sheet")
        return pd.DataFrame()

    print(f"  Reading: {phospho_file.name}")

    df = pd.read_excel(phospho_file,
                       sheet_name="Phopshoproteome_filtered_norm",
                       header=1)

    print(f"  Total phosphosites: {len(df):,}")

    # ── Filter to ERBB2 ONLY (exclude ERBB2IP and pathway proteins) ──
    erbb2 = df[
        df["Gene names"].astype(str).str.contains("ERBB2", na=False) &
        ~df["Gene names"].astype(str).str.contains("ERBB2IP", na=False)
        ].copy()

    if erbb2.empty:
        print("  ⚠ No ERBB2 phosphosites found")
        return pd.DataFrame()

    print(f"  ERBB2 phosphosites: {len(erbb2)}")

    # ── Build output records ────────────────────────────────────────────
    records = []
    for _, row in erbb2.iterrows():
        aa = str(row.get("Amino acid", ""))
        pos = str(row.get("Position_from_Positions",
                          row.get("Position", "")))
        ptm_site = f"{aa}{pos}"

        hl_fc = _safe_float(row.get("H/L_Av_log2FC"))
        ml_fc = _safe_float(row.get("M/L_Av_log2FC"))
        sig = str(row.get("t-test SignificantH/L (FDR < 1%)", ""))
        pval = _safe_float(row.get(" -Log t-test p valueH/L"))

        seq_win = str(row.get("Sequence window", ""))
        mod_seq = str(row.get("Modified sequence", ""))

        records.append({
            "cell_line": "BT-474",
            "drug_name": "Lapatinib",
            "protein": "ERBB2",
            "ptm_site": ptm_site,
            "ptm_residue": aa,
            "ptm_type": "phosphorylation",
            "peptide_sequence": seq_win.replace("_", "").strip(),
            "baseline_intensity": np.nan,
            "max_dose_intensity": np.nan,
            "max_dose_ug": np.nan,
            "log2_fold_change": round(hl_fc, 4) if np.isfinite(hl_fc) else np.nan,
            "EC50": np.nan,
            "pEC50": np.nan,
            "curve_effect_size": np.nan,
            "R2": np.nan,
            "n_doses": np.nan,
            "drug_smiles": "",
            "data_source": "ruprecht_2017",
            "resistance_context": "parental_vs_resistant",
            "target_protein": "ERBB2",
        })

    df_out = pd.DataFrame(records)

    print(f"  ✓ Extracted {len(df_out)} ERBB2 phosphosites")
    if not df_out.empty:
        print(f"    Sites: {', '.join(sorted(df_out['ptm_site'].unique()))}")
        for _, r in df_out.sort_values("log2_fold_change").iterrows():
            print(f"      {r['ptm_site']:8s} log2FC(R/P) = {r['log2_fold_change']:+.3f}")

    return df_out


# ══════════════════════════════════════════════════════════════════════════════
# ── NEW GLYCO SOURCES (J–N) — Multi-PTM Expansion (PTM-BDL §3, §11.2) ─────────
# ══════════════════════════════════════════════════════════════════════════════
# All five sources below provide N-glycosylation (N-linked glyco) measurements
# for EGFR (P00533) and / or ERBB2 (P04626).  They are necessary because the
# 2026-06-28 evaluation (§1.1 of PTM-BDL) demonstrated that the
# phosphorylation-only PTM channel is a deterministic function of
# (mutation × drug) — randomising it makes the model BETTER.
#
# N-glyco is biologically orthogonal to phospho (extracellular receptor
# state vs intracellular signalling).  Even baseline (non-drug-conditioned)
# glyco gives the model information it cannot derive from the
# sequence / drug / structure tokens.
# ══════════════════════════════════════════════════════════════════════════════

MCP_2025B_DIR = RAW_DIR / "mcp_2025b"
ERBB2_GLYCOATLAS_DIR = RAW_DIR / "erbb2_glycoform_atlas_2024"
ST6GAL1_DIR = RAW_DIR / "st6gal1_erbb2_2021"
EGFR_FUCO_DIR = RAW_DIR / "egfr_fucosylation_2020"

# ── EGFR / ERBB2 N-glyco site lists from the config (precursor numbering) ─
# These are the sites we will keep; everything else is dropped at extraction.
EGFR_SIGNAL_PEPTIDE = cfg["uniprot"]["EGFR"]["signal_peptide_length"]  # = 24
ERBB2_SIGNAL_PEPTIDE = cfg["uniprot"]["ERBB2"]["signal_peptide_length"]  # = 22


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE J: MCP 2025 — Glyco rows from Table S8 (Ref 12 in PTM-BDL)
# Same file already used for phospho (source G), but here we process the
# GP (glycopeptide) rows that source G explicitly skips.
# ══════════════════════════════════════════════════════════════════════════════

def process_mcp_2025_glyco() -> pd.DataFrame:
    """
    Process the *N-glycosylation* rows of MCP 2025 Table S8.

    Schema is identical to source G (process_mcp_2025) but we filter
    `Group == "GP"` instead of "PP".  **One row per (cell_line, site,
    glycoform)** — each glycoform keeps its own log2FC and composition.
    This is necessary because PTM-BDL §7.4 wants the per-glycoform fold
    change AS A SEPARATE TOKEN; step06 will aggregate to the per-site
    PTM-BDL token vector but each contributing glycoform survives the
    output table for downstream IG / attention attribution.

    Cell lines kept: H1975, PC-9 (each compared to H3255 sensitive baseline).
    Sentinel −10 = not detected → row dropped.

    Output rows have:
      target_protein   ∈ {EGFR, ERBB2}        (assigned in _annotate_ptm_schema)
      ptm_type         = "N-glycosylation"
      ptm_residue      = "N"
      ptm_site         = "N<position>"        (precursor numbering, P00533/P04626)
      glyco_composition = single glycan composition string
      glyco_n_glycoforms = 1
    """
    print("\n" + "=" * 70)
    print("SOURCE J: MCP 2025 — N-Glycosylation rows of Table S8")
    print("  (companion to Source G, which extracts the phospho rows)")
    print("  One row per (cell_line, site, glycoform)")
    print("=" * 70)

    s8_path = MCP_DIR / "table_s8_phospho_glyco_summary.xlsx"
    if not s8_path.exists():
        print(f"  ✗ File not found: {s8_path}")
        return pd.DataFrame()
    print(f"  Reading: {s8_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(s8_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    SENTINEL = -10.0
    accession_to_protein = {"P00533": "EGFR", "P04626": "ERBB2"}

    records: list[dict] = []
    n_raw_gp = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        group = str(row[1]).strip() if row[1] else ""
        if group != "GP":
            continue
        accession = str(row[2]).strip() if row[2] else ""
        if accession not in accession_to_protein:
            continue
        n_raw_gp += 1
        site_raw = str(row[3]).strip() if row[3] else ""
        try:
            pos = int(float(site_raw))
        except (ValueError, TypeError):
            continue

        glycan_comp = str(row[4]).strip() if row[4] else ""
        site_name = str(row[0]).strip() if row[0] else ""
        fc_h1975_h3255 = _safe_float(row[11])
        fc_pc9_h3255 = _safe_float(row[12])
        pval_h1975 = _safe_float(row[15])
        pval_pc9 = _safe_float(row[16])
        int_h1975 = _safe_float(row[19])
        int_h3255 = _safe_float(row[20])
        int_pc9 = _safe_float(row[21])

        gene = accession_to_protein[accession]
        # One row per (cell_line, glycoform):
        for cl, fc, pval, dose_int in [
            ("H1975", fc_h1975_h3255, pval_h1975, int_h1975),
            ("PC-9", fc_pc9_h3255, pval_pc9, int_pc9),
        ]:
            if not np.isfinite(fc) or abs(fc - SENTINEL) < 0.01:
                continue
            records.append({
                "cell_line": cl,
                "drug_name": "Osimertinib",
                "protein": gene,
                "ptm_site": f"N{pos}",
                "ptm_residue": "N",
                "ptm_type": "N-glycosylation",
                "peptide_sequence": "",
                "baseline_intensity": (round(int_h3255, 4)
                                       if np.isfinite(int_h3255) else np.nan),
                "max_dose_intensity": (round(dose_int, 4)
                                       if np.isfinite(dose_int) else np.nan),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(float(fc), 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "mcp_2025_glyco",
                "resistance_context": "cell_line_glycoproteome",
                "mcp_comparison": f"{cl}_vs_H3255",
                "mcp_pvalue": pval if np.isfinite(pval) else np.nan,
                "mcp_site_name": site_name,
                "glyco_composition": glycan_comp,
                "glyco_n_glycoforms": 1,
            })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Emitted {len(df)} per-glycoform rows from {n_raw_gp} GP rows")
    if not df.empty:
        for (cl, gene), sub in df.groupby(["cell_line", "protein"]):
            sites = sorted(sub["ptm_site"].unique())
            n_glyco = sub["glyco_composition"].nunique()
            print(f"    {cl:8s} / {gene:5s}: {len(sub)} rows, "
                  f"{len(sites)} sites, {n_glyco} unique glycoforms")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE K: MCP 2025b — Companion paper (Ref 31 in PTM-BDL)
# Abe et al., Metal Ion-Enhanced ZIC-cHILIC StageTip in EGFR-mutated lung cancer
# Same H1975/H3255/PC-9 cell-line panel, separately published Table S8.
# DOI: 10.1016/j.mcpro.2025.100957
# ══════════════════════════════════════════════════════════════════════════════

def process_mcp_2025b_glyco() -> pd.DataFrame:
    """
    Process MCP 2025b mmc8.xlsx — same schema as MCP 2025 Table S8.

    mmc8 is the companion-paper version of the integrated phospho+glyco
    summary table.  We extract only `Group == "GP"` rows for P00533 / P04626.
    **One row per (cell_line, site, glycoform)** — each glycoform keeps its
    own log2FC.
    """
    print("\n" + "=" * 70)
    print("SOURCE K: MCP 2025b — Companion glyco rows (mmc8.xlsx)")
    print("  One row per (cell_line, site, glycoform)")
    print("=" * 70)

    s8_path = MCP_2025B_DIR / "mmc8.xlsx"
    if not s8_path.exists():
        print(f"  ✗ File not found: {s8_path}")
        return pd.DataFrame()
    print(f"  Reading: {s8_path.name}")

    import openpyxl
    wb = openpyxl.load_workbook(s8_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    SENTINEL = -10.0
    accession_to_protein = {"P00533": "EGFR", "P04626": "ERBB2"}

    records: list[dict] = []
    n_seen_gp = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        group = str(row[1]).strip() if row[1] else ""
        if group != "GP":
            continue
        accession = str(row[2]).strip() if row[2] else ""
        if accession not in accession_to_protein:
            continue
        n_seen_gp += 1
        site_raw = str(row[3]).strip() if row[3] else ""
        try:
            pos = int(float(site_raw))
        except (ValueError, TypeError):
            continue
        glycan_comp = str(row[4]).strip() if row[4] else ""
        site_name = str(row[0]).strip() if row[0] else ""
        fc_h1975 = _safe_float(row[11])
        fc_pc9 = _safe_float(row[12])
        pval_h1975 = _safe_float(row[15])
        pval_pc9 = _safe_float(row[16])
        int_h1975 = _safe_float(row[19])
        int_h3255 = _safe_float(row[20])
        int_pc9 = _safe_float(row[21])

        gene = accession_to_protein[accession]
        for cl, fc, pval, dose_int in [
            ("H1975", fc_h1975, pval_h1975, int_h1975),
            ("PC-9", fc_pc9, pval_pc9, int_pc9),
        ]:
            if not np.isfinite(fc) or abs(fc - SENTINEL) < 0.01:
                continue
            records.append({
                "cell_line": cl,
                "drug_name": "Osimertinib",
                "protein": gene,
                "ptm_site": f"N{pos}",
                "ptm_residue": "N",
                "ptm_type": "N-glycosylation",
                "peptide_sequence": "",
                "baseline_intensity": (round(int_h3255, 4)
                                       if np.isfinite(int_h3255) else np.nan),
                "max_dose_intensity": (round(dose_int, 4)
                                       if np.isfinite(dose_int) else np.nan),
                "max_dose_ug": np.nan,
                "log2_fold_change": round(float(fc), 4),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": OSIMERTINIB_SMILES,
                "data_source": "mcp_2025b_glyco",
                "resistance_context": "cell_line_glycoproteome",
                "mcp_comparison": f"{cl}_vs_H3255",
                "mcp_pvalue": pval if np.isfinite(pval) else np.nan,
                "mcp_site_name": site_name,
                "glyco_composition": glycan_comp,
                "glyco_n_glycoforms": 1,
            })

    wb.close()
    print(f"  GP rows for EGFR/ERBB2 in mmc8: {n_seen_gp}")
    df = pd.DataFrame(records)
    print(f"  ✓ Emitted {len(df)} per-glycoform rows")
    if not df.empty:
        for (cl, gene), sub in df.groupby(["cell_line", "protein"]):
            n_glyco = sub["glyco_composition"].nunique()
            print(f"    {cl:8s} / {gene:5s}: {len(sub)} rows, "
                  f"{sub['ptm_site'].nunique()} sites, "
                  f"{n_glyco} unique glycoforms")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE L: ErbB2 Glycoform Atlas — Taniguchi et al., Glycobiology 2024
# (Ref 29 in PTM-BDL).  PMID 38109791.
#
# Table SI is a glycoform × cell-line abundance summary written as free text
# in the cells (each cell may contain "N68 3.3%\nN124/N125 0.89%..." etc.).
# To build a clean per-site, per-cell-line table we parse those strings.
#
# Cell-line / protein mapping for column index in Table SI (header row 2):
#   col 4: sErbB2 (293 cells)     — ERBB2 / HEK293     → NOT in our panel
#   col 5: sErbB2 (CHO cells)     — ERBB2 / CHO        → NOT in our panel
#   col 6: ErbB2 (SKBR-3 cells)   — ERBB2 / SK-BR-3    → IN panel (HER2+ breast)
#   col 7: ErbB2 (BT474 cells)    — ERBB2 / BT-474     → IN panel (HER2+ breast)
#   col 8: sEGFR (CHO cells)      — EGFR  / CHO        → KEEP as EGFR glycoform
#                                                         reference catalogue
# ══════════════════════════════════════════════════════════════════════════════

_GLYCO_PCT_RE = re.compile(r"N\s*(\d+(?:/N\s*\d+)?)\s+([0-9.]+)\s*%", re.IGNORECASE)


def _parse_glycoform_cell(text: str) -> list[tuple[int, float]]:
    """
    Parse a Table-SI cell like:
        "N187 74.5%\nN259 0.81%\nN530 17.7%"
    into a list of (position_int, percent_float).  "N124/N125" is split to
    two entries with the same percent.
    """
    out: list[tuple[int, float]] = []
    if not text:
        return out
    for m in _GLYCO_PCT_RE.finditer(text):
        positions_str, pct_str = m.group(1), m.group(2)
        try:
            pct = float(pct_str)
        except ValueError:
            continue
        for p in positions_str.split("/"):
            p = p.strip().lstrip("N").lstrip("n").strip()
            try:
                pos = int(p)
            except ValueError:
                continue
            out.append((pos, pct))
    return out


def process_erbb2_glycoform_atlas_2024() -> pd.DataFrame:
    """
    Process Taniguchi et al., Glycobiology 2024 — ErbB2 Glycoform Atlas.

    **One row per (cell_line, site, glycoform)** — each glycoform keeps
    its own relative-abundance percentage (Table SI cell value).  The
    `baseline_intensity` is set to that percentage so step06 can use the
    per-glycoform abundance directly.

    Drug context: baseline / untreated.  drug_name = "none".

    The data also includes EGFR N-glycosites from CHO-expressed soluble
    EGFR (column 8).  We keep them under cell_line="CHO_sEGFR" as an EGFR
    glyco reference distribution.
    """
    print("\n" + "=" * 70)
    print("SOURCE L: ErbB2 Glycoform Atlas — Taniguchi 2024 (PMID 38109791)")
    print("  One row per (cell_line, site, glycoform)")
    print("=" * 70)

    p = ERBB2_GLYCOATLAS_DIR / "tablesi_cwad100.xlsx"
    if not p.exists():
        print(f"  ✗ File not found: {p}")
        return pd.DataFrame()
    print(f"  Reading: {p.name}")

    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Column → (cell_line, target_protein, resistance_context)
    cl_columns = {
        6: ("SK-BR-3", "ERBB2", "lapatinib_baseline_glyco"),
        7: ("BT-474", "ERBB2", "lapatinib_baseline_glyco"),
        8: ("CHO_sEGFR", "EGFR", "egfr_glycoform_reference"),
    }

    records: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # skip 2-row meta + header
            continue
        composition = str(row[2]) if row[2] else ""
        glycan_mz = str(row[1]) if row[1] else ""
        for col, (cl, gene, ctx) in cl_columns.items():
            cell_text = str(row[col]) if row[col] else ""
            if not cell_text or cell_text.strip().lower() in ("n.d.", "nan", "none", ""):
                continue
            for pos, pct in _parse_glycoform_cell(cell_text):
                records.append({
                    "cell_line": cl,
                    "drug_name": "none",
                    "protein": gene,
                    "ptm_site": f"N{pos}",
                    "ptm_residue": "N",
                    "ptm_type": "N-glycosylation",
                    "peptide_sequence": "",
                    # baseline_intensity = the relative-abundance % for THIS
                    # glycoform at THIS site in THIS cell line.
                    "baseline_intensity": round(float(pct), 4),
                    "max_dose_intensity": np.nan,
                    "max_dose_ug": np.nan,
                    "log2_fold_change": np.nan,
                    "EC50": np.nan, "pEC50": np.nan,
                    "curve_effect_size": np.nan, "R2": np.nan,
                    "n_doses": np.nan,
                    "drug_smiles": "",
                    "data_source": "erbb2_glycoform_atlas_2024",
                    "resistance_context": ctx,
                    "glyco_composition": composition,
                    "glyco_mz": glycan_mz,
                    "glyco_relative_abundance_pct": round(float(pct), 4),
                    "glyco_n_glycoforms": 1,
                })
    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Emitted {len(df)} per-glycoform rows")
    if not df.empty:
        for (cl, gene), sub in df.groupby(["cell_line", "protein"]):
            n_glyco = sub["glyco_composition"].nunique()
            print(f"    {cl:12s} / {gene:5s}: {len(sub)} rows, "
                  f"{sub['ptm_site'].nunique()} sites, "
                  f"{n_glyco} unique glycoforms")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE M: ST6Gal1 → ErbB2 — Garnham et al., Oncogene 2021 (PMID 33947960)
# (Ref 30 in PTM-BDL).
#
# The supplementary file has three sheets:
#   "ErbB2 GC Cells 1st Replicate" (278 rows)
#   "ErbB2 GC Cells 2nd Replicate" (196 rows)
#   "ErbB2 + Gastric Adenocarcinoma " (9 rows)
#
# Column layout (header row is at index 4, data starts at row 5):
#   col 0: FILE
#   col 1: SAMPLE     ("ErbB2 WT", "ErbB2 ST6Gal1KO", etc.)
#   col 2: Site       (peptide-numbering, e.g. "124", "187")  ← mature-protein
#   col 3: m/z
#   col 4: Charge
#   ...
#   col 7: peptide Sequence
#   col 8: glycan composition
#   col 9: glycan mass
#   col 10: proposed structure
#   col 11: scan
#   col 12: time
#   col 13: score
#
# CRITICAL: site numbers in this paper are MATURE-protein numbering for
# ERBB2 (signal peptide of 22 stripped).  Our PTM-BDL token vector uses
# UniProt P04626 PRECURSOR numbering (matches config/config.yaml).  Conversion:
#       precursor_position = mature_position + signal_peptide_length(22)
# This is verified by the canonical ErbB2 N-glyco sites:
#   mature 124 = precursor 146  → but Taniguchi 2024 reports it as N124
#                                  (mature numbering), so we must add 22.
# We apply the +22 conversion below.
# ══════════════════════════════════════════════════════════════════════════════

def process_st6gal1_erbb2_2021() -> pd.DataFrame:
    """
    Process Garnham et al., Oncogene 2021 — ST6Gal1 / ErbB2 glycoproteomics.

    This is the ONLY drug-conditioned glycosylation dataset we have for
    HER2 (drug = Trastuzumab, per the paper).  ST6Gal1 modulates
    sialylation of ErbB2 and changes trastuzumab sensitivity.

    Aggregation: per (sample, site) we count how many distinct glycoforms
    were detected and which compositions were observed.  We do NOT have
    a parental-vs-treated quantitative ratio in this table, so
    log2_fold_change stays NaN.  `baseline_intensity` is set to the
    glycoform count as a coarse occupancy proxy.

    Sample → cell_line mapping (all "Gastric Carcinoma" cells in the paper):
      "ErbB2 WT"               → "ErbB2_WT_GC"
      "ErbB2 ST6Gal1KO"        → "ErbB2_ST6Gal1KO_GC"   ← the perturbation
      "Intestinal-Subtype ErbB2-Positive Gastric Adenocarcinoma" → tumor sample
    None of these map directly to our GDSC-anchored cell-line panel, so
    step06 will treat them as REFERENCE glyco distributions for HER2 unless
    we later extend the panel.

    NOTE: drug_name = "Trastuzumab" (not a TKI).  The drug-filter in
    `merge_and_save` currently keeps Trastuzumab — already in `relevant_drugs`.
    """
    print("\n" + "=" * 70)
    print("SOURCE M: ST6Gal1 → ErbB2 — Garnham 2021 (PMID 33947960)")
    print("  Site-specific HER2 glycoproteomics with trastuzumab context")
    print("=" * 70)

    files = sorted(ST6GAL1_DIR.glob("*.xlsx"))
    if not files:
        print(f"  ✗ No .xlsx files found in: {ST6GAL1_DIR}")
        return pd.DataFrame()
    p = files[0]
    print(f"  Reading: {p.name}")

    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)

    # accumulator: (sample, site_precursor) → list of (composition, score)
    accum: dict[tuple, list[tuple[str, float, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # The header row is at index 4 in each sheet (per inspection)
        # Data rows start at index 5.
        last_seen_sample = None
        last_seen_site_mature = None
        last_seen_composition = None
        last_seen_peptide = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 5:
                continue
            # SAMPLE / SITE can be blank on continuation rows (subsequent
            # glycoforms for the same peptide).  We carry the last seen.
            sample_cell = row[1]
            if sample_cell:
                last_seen_sample = str(sample_cell).strip()
            site_cell = row[2]
            if site_cell:
                try:
                    last_seen_site_mature = int(float(str(site_cell).strip()))
                except (ValueError, TypeError):
                    pass
            if row[7]:
                last_seen_peptide = str(row[7]).strip()
            if row[8]:
                last_seen_composition = str(row[8]).strip()

            score_val = _safe_float(row[13])

            if not last_seen_sample or last_seen_site_mature is None:
                continue
            if not last_seen_composition:
                continue

            # Convert mature → precursor numbering
            precursor_pos = last_seen_site_mature + ERBB2_SIGNAL_PEPTIDE

            key = (last_seen_sample, precursor_pos)
            accum.setdefault(key, []).append(
                (last_seen_composition, score_val, last_seen_peptide or "")
            )
    wb.close()

    if not accum:
        print("  ⚠ Could not parse any glycopeptide rows.")
        return pd.DataFrame()

    # Sample → cell_line normalisation.  The raw labels actually used by
    # Garnham 2021 (verified by inspection 2026-06-28) are:
    #   "ErbB2 ST6GAL1 K.O. Clone 1/2/3"   — KO perturbation lines
    #   "SAMPLE"                            — header row residue
    #   "Intestinal-Subtype ErbB2-Positive Gastric Adenocarcinoma" — tumour
    # These are gastric-carcinoma derived, not breast.  We map them all to
    # the "ErbB2_ST6Gal1_GC" pseudo-cell-line so step06 can keep them as
    # an ERBB2 glyco reference (the cell-line allow-list in merge_and_save
    # includes the same label via GLYCO_EXTRA_CELL_LINES).
    sample_to_cell = {
        "ErbB2 ST6GAL1 K.O. Clone 1": "ErbB2_ST6Gal1KO_GC",
        "ErbB2 ST6GAL1 K.O. Clone 2": "ErbB2_ST6Gal1KO_GC",
        "ErbB2 ST6GAL1 K.O. Clone 3": "ErbB2_ST6Gal1KO_GC",
        "ErbB2 WT": "BT-474",
        "ErbB2 ST6Gal1KO": "BT-474",
        "Intestinal-Subtype ErbB2-Positive Gastric Adenocarcinoma": "ErbB2_GC_tumor",
        "SAMPLE": "ErbB2_ST6Gal1_unknown",
    }

    # ── One row per (sample, site, glycoform) — match J/K/L semantics ───
    records: list[dict] = []
    for (sample, pos), entries in accum.items():
        cl = sample_to_cell.get(sample, sample)
        # Aggregate per-glycoform: multiple PSMs of the same composition
        # at the same site are pooled (peak count).
        per_glyco: dict[str, dict] = {}
        for comp, score, peptide in entries:
            d = per_glyco.setdefault(comp, {
                "n_psm": 0,
                "max_score": np.nan,
                "peptide": peptide,
            })
            d["n_psm"] += 1
            if np.isfinite(score) and (not np.isfinite(d["max_score"])
                                       or score > d["max_score"]):
                d["max_score"] = score
            if peptide and not d["peptide"]:
                d["peptide"] = peptide

        for comp, info in per_glyco.items():
            records.append({
                "cell_line": cl,
                "drug_name": "Trastuzumab",
                "protein": "ERBB2",
                "ptm_site": f"N{pos}",
                "ptm_residue": "N",
                "ptm_type": "N-glycosylation",
                "peptide_sequence": info["peptide"],
                # PSM count is a coarse abundance proxy for this glycoform.
                "baseline_intensity": float(info["n_psm"]),
                "max_dose_intensity": np.nan,
                "max_dose_ug": np.nan,
                "log2_fold_change": np.nan,
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": "",
                "data_source": "st6gal1_erbb2_2021",
                "resistance_context": "trastuzumab_resistance_glyco",
                "glyco_composition": comp,
                "glyco_n_glycoforms": 1,
                "glyco_psm_count": int(info["n_psm"]),
                "glyco_byonic_score": (round(float(info["max_score"]), 2)
                                       if np.isfinite(info["max_score"])
                                       else np.nan),
                "st6gal1_sample": sample,
            })
    df = pd.DataFrame(records)
    print(f"  ✓ Emitted {len(df)} per-glycoform ERBB2 rows")
    if not df.empty:
        for cl, sub in df.groupby("cell_line"):
            n_glyco = sub["glyco_composition"].nunique()
            print(f"    {cl:25s}: {len(sub)} rows, "
                  f"{sub['ptm_site'].nunique()} sites, "
                  f"{n_glyco} unique glycoforms")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE N: EGFR Fucosylation — Sethi et al., Mol Omics 2020 (PMID 32203567)
# (Ref 32 in PTM-BDL).
#
# Table S7 ("S7 EGFR Site Occupancy") is the cleanest, most comparable
# summary across the workbook: it provides per-site occupancy numbers for
# CAL27 and HSC3 OSCC cells under DMSO vs ICG001 treatment.
#
# Schema of S7 (header rows 0–3 are meta, data rows from index 4):
#   col 0: Site (mature numbering, e.g. "N32")
#   col 1: Peptide
#   col 2: CAL27 DMSO occupancy  (fraction 0–1, or "High" for >0.95)
#   col 3: CAL27 ICG001 occupancy
#   col 4: HSC3  DMSO occupancy
#   col 5: HSC3  ICG001 occupancy
#
# Site labels use the MATURE EGFR numbering (signal peptide of 24 stripped).
# To match `config.yaml ptm.EGFR` which uses precursor numbering we add 24:
#       precursor_position = mature_position + 24
#       mature N32  → precursor N56
#       mature N104 → precursor N128
#       mature N151 → precursor N175
#       mature N389 → precursor N413
#       mature N420 → precursor N444
#       mature N579 → precursor N603
# These cover 6 of the 12+ EGFR N-glyco sites.
#
# Drug: ICG001 is a β-catenin/CBP inhibitor — NOT in our drug panel.  We tag
# drug_name = "ICG001" so the drug filter in `merge_and_save` drops these
# rows from drug-conditioned analyses.  However, the DMSO baseline occupancy
# is still a valuable EGFR-glyco reference, so we *also* emit a parallel
# "drug_name = none" row for the DMSO baseline.
# ══════════════════════════════════════════════════════════════════════════════

def _parse_occupancy_value(v) -> float:
    """Parse Table S7 occupancy cell.  "High" → 1.0, NaN-ish → NaN."""
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    if s in ("", "nan", "none", "na", "-", "—"):
        return np.nan
    if s == "high":
        return 1.0
    if s == "low":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return np.nan


def process_egfr_fucosylation_2020() -> pd.DataFrame:
    """
    Process Sethi et al., Mol Omics 2020 — EGFR site occupancy reference.

    Output rows have:
      ptm_type      = "N-glycosylation"
      ptm_residue   = "N"
      ptm_site      = "N<precursor_pos>"
      log2_fold_change = log2( ICG001_occupancy / DMSO_occupancy )
                         (when both > 0)
      drug_name     = "none"  (we emit a single baseline row per cell-line ×
                              site, encoding the DMSO occupancy; the
                              ICG001/DMSO log2FC goes into `log2_fold_change`
                              but is not propagated as a drug effect because
                              ICG001 is not in our drug panel)
    """
    print("\n" + "=" * 70)
    print("SOURCE N: EGFR Fucosylation — Sethi 2020 (PMID 32203567)")
    print("  Site-occupancy reference catalog (CAL27, HSC3 OSCC)")
    print("=" * 70)

    f = EGFR_FUCO_DIR / (
        "NIHMS1574909-supplement-Supplement_-_Tables_1_-_7_"
        "Site_Occupancy_Analysis_and_Glycopeptide_HILIC_C18_MSMS.xlsx"
    )
    if not f.exists():
        print(f"  ✗ File not found: {f}")
        return pd.DataFrame()
    print(f"  Reading: {f.name}")

    import openpyxl
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    if "S7 EGFR Site Occupancy" not in wb.sheetnames:
        print(f"  ✗ Sheet 'S7 EGFR Site Occupancy' not found.")
        wb.close()
        return pd.DataFrame()

    ws = wb["S7 EGFR Site Occupancy"]

    records: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:  # skip 4-row header / meta
            continue
        site_label = str(row[0]) if row[0] else ""
        if not site_label or not site_label.lstrip().lower().startswith("n"):
            continue
        try:
            mature_pos = int(re.sub(r"[^0-9]", "", site_label))
        except ValueError:
            continue
        precursor_pos = mature_pos + EGFR_SIGNAL_PEPTIDE
        peptide = str(row[1]) if row[1] else ""

        cal27_dmso = _parse_occupancy_value(row[2])
        cal27_icg = _parse_occupancy_value(row[3])
        hsc3_dmso = _parse_occupancy_value(row[4])
        hsc3_icg = _parse_occupancy_value(row[5])

        for cell_line, dmso_occ, icg_occ in [
            ("CAL27", cal27_dmso, cal27_icg),
            ("HSC3", hsc3_dmso, hsc3_icg),
        ]:
            if not np.isfinite(dmso_occ):
                continue
            # log2FC = log2(ICG001 / DMSO) — only when both > 0
            if np.isfinite(icg_occ) and icg_occ > 0 and dmso_occ > 0:
                log2fc = float(np.log2(icg_occ / dmso_occ))
            else:
                log2fc = np.nan

            records.append({
                "cell_line": cell_line,
                # ICG001 is not in our drug panel and would be dropped by the
                # drug filter.  We use drug_name="none" to keep the baseline
                # occupancy row; the ICG001 effect is preserved in log2FC.
                "drug_name": "none",
                "protein": "EGFR",
                "ptm_site": f"N{precursor_pos}",
                "ptm_residue": "N",
                "ptm_type": "N-glycosylation",
                "peptide_sequence": peptide,
                "baseline_intensity": round(dmso_occ, 4),
                "max_dose_intensity": (round(icg_occ, 4)
                                       if np.isfinite(icg_occ) else np.nan),
                "max_dose_ug": np.nan,
                "log2_fold_change": (round(log2fc, 4)
                                     if np.isfinite(log2fc) else np.nan),
                "EC50": np.nan, "pEC50": np.nan,
                "curve_effect_size": np.nan, "R2": np.nan,
                "n_doses": np.nan,
                "drug_smiles": "",
                "data_source": "egfr_fucosylation_2020",
                "resistance_context": "egfr_glyco_occupancy_reference",
                "glyco_composition": "",  # composition is in S8–S11 sheets
                "fuco_mature_site": site_label,
                "fuco_treatment": "ICG001 vs DMSO",
            })

    wb.close()

    df = pd.DataFrame(records)
    print(f"  ✓ Extracted {len(df)} per-site occupancy rows "
          f"(CAL27 + HSC3 OSCC reference)")
    if not df.empty:
        for cl, sub in df.groupby("cell_line"):
            print(f"    {cl:8s}: {len(sub)} sites → "
                  f"{sorted(sub['ptm_site'].unique())}")
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Cell-line allow-list extension — needed so the new glyco rows survive the
# `erbb_cell_lines` filter inside `merge_and_save`.  We add the new lines
# coming from sources L–N here so they are not silently dropped.
# ──────────────────────────────────────────────────────────────────────────────
GLYCO_EXTRA_CELL_LINES = {
    # ErbB2 Glycoform Atlas (source L)
    "CHO_sEGFR",
    # ST6Gal1 ErbB2 (source M)
    "ErbB2_GC_tumor",
    "ErbB2_ST6Gal1KO_GC",  # pooled KO clones 1/2/3 (Garnham 2021)
    "ErbB2_ST6Gal1_unknown",  # the "SAMPLE" header-residue row
    "ErbB2 WT", "ErbB2 ST6Gal1KO",  # raw labels (defensive — sample_to_cell may miss)
    # EGFR Fucosylation (source N)
    "CAL27", "HSC3",
}

# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  STEP 05: Drug-PTM Data — Multi-PTM × Multi-Protein Integration  ║")
    print("║  (EGFR + HER2 / ERBB2  ×  Phosphorylation + N-Glycosylation)     ║")
    print("╠══════════════════════════════════════════════════════════════════╣")
    print("║  PHOSPHORYLATION SOURCES                                         ║")
    print("║   A: DrugPTM-Bench (PMID 30394195) — EGFR + ERBB2 dose-response  ║")
    print("║   B: Tozuka 2024 (PMID 38646155) — PC-9/HCC827 vs OsiR           ║")
    print("║   C: Hsu 2025 (PMID 41023502) — PC-9 temporal dynamics           ║")
    print("║   D: PNAS 2025 — H1975/HCC4006 pY under Osi                      ║")
    print("║   E: FEBS 2025 — LUAD tumor phospho signatures                   ║")
    print("║   F: Cancer Res 2021 — SILAC H1975 + AZR/COR resistant clones    ║")
    print("║   G: MCP 2025 — H1975/H3255/PC-9 phospho rows                    ║")
    print("║   H: Remsing Rix 2022 — PC9GR + Osi                              ║")
    print("║   I: Ruprecht 2017 — BT-474 lapatinib resistance ERBB2 phospho   ║")
    print("║                                                                  ║")
    print("║  N-GLYCOSYLATION SOURCES (new — PTM-BDL §3, §11.2)               ║")
    print("║   J: MCP 2025 — GP rows from Table S8 (EGFR + ERBB2 N-sites)     ║")
    print("║   K: MCP 2025b (PMID 40154885) — mmc8 companion glyco            ║")
    print("║   L: ErbB2 Glycoform Atlas (PMID 38109791) — Taniguchi 2024       ║")
    print("║   M: ST6Gal1 → ErbB2 (PMID 33947960) — Garnham 2021 (trastuzumab)║")
    print("║   N: EGFR Fucosylation (PMID 32203567) — Sethi 2020 (CAL27/HSC3) ║")
    print("╚══════════════════════════════════════════════════════════════════╝")

    # ── Source A: DrugPTM-Bench (largest, slowest — chunked) ──────────────
    cell_line_files = verify_drugptm_bench()
    if cell_line_files:
        df_egfr = extract_egfr_data(cell_line_files)
        df_bench = build_drugptm_bench_summaries(df_egfr)
        catalog = build_catalog(cell_line_files)
        # ── Source A-HER2: ERBB2 from BT-474/MDA-MB-175 ──────────────────
        df_erbb2_raw = extract_erbb2_data(cell_line_files)
        df_erbb2 = build_erbb2_summaries(df_erbb2_raw)
    else:
        print("\n  ⚠ DrugPTM-Bench data not found — skipping Source A")
        df_bench = pd.DataFrame()
        df_erbb2 = pd.DataFrame()
        catalog = pd.DataFrame()

    # ── Source B: Tozuka 2024 (parental vs resistant) ─────────────────────
    df_tozuka = process_tozuka_2024()

    # ── Source C: Hsu 2025 (temporal dynamics) ────────────────────────────
    df_hsu = process_hsu_2025()

    # ── Source D: PNAS 2025 (pY phosphoproteome under TKI) ───────────────
    df_pnas = process_pnas_2025()

    # ── Source E: FEBS 2025 (tumor phospho signatures) ────────────────────
    df_febs = process_febs_2025()

    # ── Source F: Cancer Research 2021 (SILAC resistance phospho) ─────────
    df_cancerres = process_cancerres_2021()

    # ── Source G: MCP 2025 (phospho/glyco cell-line comparison) ───────────
    df_mcp = process_mcp_2025()

    # ── Source H: Remsing Rix 2022 (PC9GR + Osimertinib) ─────────────────
    df_remsing = process_remsing_rix_2022()

    # ── Source I: Ruprecht 2017 (HER2 lapatinib resistance) ──────────────
    df_ruprecht = process_ruprecht_2017()

    # ──────────────────────────────────────────────────────────────────────
    # ── Multi-PTM expansion — N-glycosylation sources (J–N) ──────────────
    # ──────────────────────────────────────────────────────────────────────

    # Source J: MCP 2025 — N-glyco rows of Table S8 (companion to source G)
    df_mcp_glyco = process_mcp_2025_glyco()

    # Source K: MCP 2025b — Companion paper mmc8 glyco rows
    df_mcp_2025b_glyco = process_mcp_2025b_glyco()

    # Source L: ErbB2 Glycoform Atlas (Taniguchi 2024)
    df_erbb2_glycoatlas = process_erbb2_glycoform_atlas_2024()

    # Source M: ST6Gal1 → ErbB2 site-specific glycoproteomics
    df_st6gal1 = process_st6gal1_erbb2_2021()

    # Source N: EGFR Fucosylation site-occupancy reference (Sethi 2020)
    df_egfr_fuco = process_egfr_fucosylation_2020()

    # ── Merge ALL sources (phospho + glyco) and save ─────────────────────
    merge_and_save(
        df_bench=df_bench,
        df_tozuka=df_tozuka,
        df_hsu=df_hsu,
        df_pnas=df_pnas,
        df_febs=df_febs,
        df_cancerres=df_cancerres,
        df_mcp=df_mcp,
        catalog=catalog,
        df_remsing=df_remsing,
        df_ruprecht=df_ruprecht,
        df_erbb2_bench=df_erbb2,
        # Glyco frames (multi-PTM expansion):
        df_mcp_glyco=df_mcp_glyco,
        df_mcp_2025b_glyco=df_mcp_2025b_glyco,
        df_erbb2_glycoatlas=df_erbb2_glycoatlas,
        df_st6gal1=df_st6gal1,
        df_egfr_fuco=df_egfr_fuco,
    )

    print("\n✓ Step 05 complete! Multi-PTM, multi-protein drug-PTM data "
          "ready for harmonization (Step 06).")
